from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from datetime import datetime, timedelta
from functools import wraps
import os
import io
import pandas as pd
from models import db, User, TenantCustomer, Organization, Customer, Product, Contract, ContractProduct, Payment, Delivery, Invoice, SysConfig, Position, UserPosition, PositionContractPermission, UserOrganization

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
db_url = os.environ.get('DATABASE_URL', 'sqlite:///contracts.db')
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['UPLOAD_FOLDER'] = 'uploads'
db.init_app(app)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


# ── 自定义模板过滤器：去掉浮点数尾部多余的0，如 6.0 → 6，6.5 → 6.5 ──
@app.template_filter('notrailzero')
def notrailzero_filter(val):
    if val is None:
        return ''
    return '{:g}'.format(float(val))


# 模型已移至 models.py


with app.app_context():
    db.create_all()
    # 初始化超级管理员 admin（无租户，可管理所有数据）
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', role='超级管理员', permissions='all', customer_id=None)
        admin.set_password('123456')
        db.session.add(admin)
        db.session.commit()
    # 新增：初始化 superadmin（总超级管理员，可创建客户超级管理员）
    if not User.query.filter_by(username='superadmin').first():
        sa = User(username='superadmin', role='超级管理员', permissions='all', customer_id=None)
        sa.set_password('654321')
        db.session.add(sa)
        db.session.commit()


@app.context_processor
def inject_company():
    try:
        configs = {c.key: c.value for c in SysConfig.query.all()}
    except Exception:
        configs = {}

    # 默认使用全局配置
    company_name = configs.get('company_name', '')
    company_logo_file = configs.get('company_logo_file', '')
    system_name = configs.get('system_name', '客户管理系统')

    # 若当前登录用户属于某租户，优先用租户自己的品牌信息
    try:
        user_id = session.get('user_id')
        if user_id:
            user = db.session.get(User, user_id)
            if user and user.customer_id:
                tenant = db.session.get(TenantCustomer, user.customer_id)
                if tenant:
                    if tenant.system_name:
                        system_name = tenant.system_name
                    if tenant.company_name:
                        company_name = tenant.company_name
                    if tenant.logo_file:
                        company_logo_file = tenant.logo_file
    except Exception:
        pass

    current_tenant_name = None
    try:
        user_id = session.get('user_id')
        if user_id:
            user = db.session.get(User, user_id)
            if user and user.customer_id:
                t = db.session.get(TenantCustomer, user.customer_id)
                if t:
                    current_tenant_name = t.name
    except Exception:
        pass

    return dict(company_name=company_name, company_logo_file=company_logo_file, current_tenant_name=current_tenant_name, system_name=system_name)


# ── 辅助：判断当前登录者是否为 superadmin ──
def is_superadmin():
    return session.get('username') == 'superadmin'


# ── 新增：获取当前用户的租户ID ──
def get_current_customer_id():
    """获取当前登录用户的租户客户ID，superadmin返回None"""
    if 'user_id' not in session:
        return None
    user = db.session.get(User, session['user_id'])
    return user.customer_id if user else None


# ── 新增：判断是否为客户超级管理员 ──
def is_customer_admin():
    """判断当前用户是否为客户超级管理员（admin角色且有customer_id）"""
    if 'user_id' not in session:
        return False
    user = db.session.get(User, session['user_id'])
    return user and user.role == '超级管理员' and user.customer_id is not None


# ── 新增：功能权限计算辅助函数 ──
def get_user_function_permissions(user_id):
    """
    获取用户的综合功能权限（用户自身权限 + 岗位权限合并）
    返回：权限列表或 'all'
    """
    user = db.session.get(User, user_id)
    if not user:
        return []

    # 超级管理员拥有全部权限
    if user.role == '超级管理员' or user.permissions == 'all':
        return 'all'

    # 收集所有权限
    permissions = set()

    # 1. 用户自身的权限
    if user.permissions:
        if user.permissions == 'all':
            return 'all'
        permissions.update(user.permissions.split(','))

    # 2. 用户岗位的功能权限
    user_positions = UserPosition.query.filter_by(user_id=user_id).all()
    for up in user_positions:
        position = db.session.get(Position, up.position_id)
        if position and position.function_permissions:
            position_perms = position.function_permissions.split(',')
            permissions.update(position_perms)

    return list(permissions)


def get_user_position_permissions_info(user_id):
    """
    获取用户的岗位权限详细信息（用于用户管理页面显示）
    返回：{
        'positions': [{'name': '岗位名', 'permissions': '权限字符串', 'organization': '组织名'}],
        'merged_permissions': '合并后的权限字符串'
    }
    """
    user = db.session.get(User, user_id)
    if not user:
        return {'positions': [], 'merged_permissions': ''}

    positions_list = []
    all_perms = set()

    user_positions = UserPosition.query.filter_by(user_id=user_id).all()
    for up in user_positions:
        position = db.session.get(Position, up.position_id)
        if position:
            org_name = ''
            if up.organization_id:
                org = db.session.get(Organization, up.organization_id)
                if org:
                    org_name = org.name

            positions_list.append({
                'name': position.name,
                'permissions': position.function_permissions or '',
                'organization': org_name,
                'is_primary': up.is_primary
            })

            if position.function_permissions:
                all_perms.update(position.function_permissions.split(','))

    merged_permissions = ','.join(sorted(all_perms)) if all_perms else ''

    return {
        'positions': positions_list,
        'merged_permissions': merged_permissions
    }


def has_permission(user_id, permission):
    """
    检查用户是否拥有指定的功能权限
    参数：
        user_id: 用户ID
        permission: 权限名称（如 '增加', '删除', '修改', '查阅', '下载', '增加-收付款'）
    返回：True/False

    支持新旧权限格式：
    - 旧格式权限 '增加' 可以匹配新格式检查 '增加-合同'（向下兼容）
    - 新格式权限 '增加-合同' 不能匹配旧格式检查 '增加'（精确控制）
    """
    permissions = get_user_function_permissions(user_id)

    # 'all' 表示拥有所有权限
    if permissions == 'all':
        return True

    # 完全匹配检查
    if permission in permissions:
        return True

    # 如果检查的是细粒度权限（如 '增加-收付款'），且用户有旧格式权限（如 '增加'）
    # 则认为有权限（向下兼容旧系统）
    if '-' in permission:
        base_permission = permission.split('-')[0]  # 提取 '增加-收付款' 中的 '增加'
        if base_permission in permissions:
            return True

    return False


def has_permission_for_contract(user_id, permission, contract):
    """
    检查用户是否对特定合同拥有指定的功能权限
    项目负责人自动拥有该合同的所有权限

    参数：
        user_id: 用户ID
        permission: 权限名称（如 '增加', '删除', '修改', '查阅'）
        contract: 合同对象
    返回：True/False
    """
    user = db.session.get(User, user_id)
    if not user:
        return False

    # 检查是否是该合同的项目负责人
    if contract and contract.project_staff and user.username in contract.project_staff:
        # 项目负责人拥有该合同的所有操作权限
        return True

    # 否则检查常规权限
    return has_permission(user_id, permission)


def get_user_permissions_string(user_id):
    """
    获取用户权限的字符串表示（用于传递给模板）
    返回：'all' 或逗号分隔的权限列表
    """
    permissions = get_user_function_permissions(user_id)

    if permissions == 'all':
        return 'all'

    # 返回逗号分隔的权限字符串
    return ','.join(permissions) if permissions else ''


# ── 新增：数据权限计算辅助函数 ──
def get_organization_with_children(org_id):
    """
    获取组织及其所有下级组织的ID列表（递归）
    用于"本组织合同"权限：上级组织能看到所有下级组织的合同

    参数：
        org_id: 组织ID
    返回：
        [org_id, child1_id, child2_id, ...] 包含自身和所有后代组织
    """
    if not org_id:
        return []

    result = [org_id]

    # 查找直接子组织
    children = Organization.query.filter_by(parent_id=org_id).all()

    # 递归获取每个子组织的后代
    for child in children:
        result.extend(get_organization_with_children(child.id))

    return result


def get_user_data_scope(user_id):
    """
    计算用户的综合数据权限范围
    返回：'all' | 'org' | 'custom' | 'self'
    优先级：all > org > custom > self
    """
    user = db.session.get(User, user_id)
    if not user:
        return 'self'

    # 超级管理员拥有全部权限
    if user.role == '超级管理员' or user.permissions == 'all':
        return 'all'

    # 查询用户的所有岗位
    user_positions = UserPosition.query.filter_by(user_id=user_id).all()
    if not user_positions:
        return 'self'  # 没有岗位时，默认只能看自己创建的

    # 获取所有岗位的数据权限范围
    scopes = []
    for up in user_positions:
        position = db.session.get(Position, up.position_id)
        if position:
            scopes.append(position.data_scope)

    # 按优先级返回最宽松的权限
    if 'all' in scopes:
        return 'all'
    if 'org' in scopes:
        return 'org'
    if 'custom' in scopes:
        return 'custom'
    return 'self'


def get_user_accessible_contract_ids(user_id, customer_id, include_finished=False):
    """
    获取用户可访问的合同ID列表
    参数：
        user_id: 用户ID
        customer_id: 租户ID
        include_finished: 是否包含已完结合同（默认False，只显示进行中）
    返回：合同ID列表，如果返回 None 表示可访问所有合同
    """
    user = db.session.get(User, user_id)
    if not user:
        return []

    # 超级管理员可访问所有合同
    if user.role == '超级管理员' or user.permissions == 'all':
        return None  # None 表示全部

    data_scope = get_user_data_scope(user_id)

    # 全部合同权限
    if data_scope == 'all':
        return None  # None 表示全部

    # 本组织合同权限
    if data_scope == 'org':
        # 获取用户的主组织（只看主组织的数据）
        user_orgs = UserOrganization.query.filter_by(user_id=user_id, is_primary=True).all()
        org_ids = [uo.organization_id for uo in user_orgs]

        # 新增：已完结的合同只有原合同负责人/创建人能看到（不随组织显示）
        # 无论是否有主组织，用户都能看到自己负责/创建的合同（含已完结）
        own_contracts = Contract.query.filter(
            Contract.customer_id == customer_id,
            db.or_(
                Contract.created_by == user_id,
                Contract.project_staff.like(f'%{user.username}%')
            )
        ).all()
        own_ids = [c.id for c in own_contracts]

        # 新增：通过岗位被显式授权的合同（含已完结，原负责人及被授权者都能看到）
        user_positions = UserPosition.query.filter_by(user_id=user_id).all()
        position_ids = [up.position_id for up in user_positions]
        granted_ids = []
        if position_ids:
            permissions = PositionContractPermission.query.filter(
                PositionContractPermission.position_id.in_(position_ids)
            ).all()
            granted_ids = [p.contract_id for p in permissions]

        if not org_ids:
            # 如果用户没有主组织，则只能看自己创建/负责的和被授权的合同
            return list(set(own_ids + granted_ids))

        # 新增：扩展组织ID列表，包含所有下级组织（支持组织层级管理）
        # 例如：用户在"项目部"，能看到项目部及其下所有子项目的合同
        expanded_org_ids = []
        for org_id in org_ids:
            expanded_org_ids.extend(get_organization_with_children(org_id))
        # 去重
        expanded_org_ids = list(set(expanded_org_ids))

        # 本组织合同：默认只显示"进行中"的合同，用户可选包含已完结
        # 查询条件：本组织及所有下级组织的合同
        query = Contract.query.filter(
            Contract.customer_id == customer_id,
            Contract.organization_id.in_(expanded_org_ids)
        )

        # 根据 include_finished 参数决定是否过滤已完结合同
        if not include_finished:
            query = query.filter(Contract.status != '已完结')

        org_contracts = query.all()
        org_ids_list = [c.id for c in org_contracts]

        # 合并：本组织及下级组织的合同 + 自己负责/创建的合同（含已完结）+ 被岗位授权的合同（含已完结）
        return list(set(org_ids_list + own_ids + granted_ids))

    # 自定义权限
    if data_scope == 'custom':
        # 获取用户所有岗位被授权的合同
        user_positions = UserPosition.query.filter_by(user_id=user_id).all()
        position_ids = [up.position_id for up in user_positions]

        # 查询这些岗位被授权的合同
        permissions = PositionContractPermission.query.filter(
            PositionContractPermission.position_id.in_(position_ids)
        ).all()

        contract_ids = list(set([p.contract_id for p in permissions]))

        # 自定义权限的用户也能看到自己创建的和负责的合同
        own_contracts = Contract.query.filter(
            Contract.customer_id == customer_id,
            db.or_(
                Contract.created_by == user_id,
                Contract.project_staff.like(f'%{user.username}%')
            )
        ).all()
        own_contract_ids = [c.id for c in own_contracts]

        return list(set(contract_ids + own_contract_ids))

    # 仅自己创建的合同和负责的合同
    own_contracts = Contract.query.filter(
        Contract.customer_id == customer_id,
        db.or_(
            Contract.created_by == user_id,
            Contract.project_staff.like(f'%{user.username}%')
        )
    ).all()
    return [c.id for c in own_contracts]


# 登录验证装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# 权限验证装饰器
def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('请先登录', 'warning')
                return redirect(url_for('login'))

            user_id = session['user_id']

            # 使用新的权限检查函数（整合用户自身权限 + 岗位权限）
            if has_permission(user_id, permission):
                return f(*args, **kwargs)

            flash('你没有此项权限，请与管理员联系', 'warning')
            return redirect(url_for('index'))
        return decorated_function
    return decorator


@app.route('/login', methods=['GET', 'POST'])
def login():
    # 获取系统配置
    try:
        configs = {c.key: c.value for c in SysConfig.query.all()}
        system_name = configs.get('system_name', '客户管理系统')
    except Exception:
        system_name = '客户管理系统'
    tenants = TenantCustomer.query.order_by(TenantCustomer.name).all()
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        customer_id = request.form.get('customer_id', '').strip()
        customer_name = request.form.get('customer_name', '').strip()

        # 如果提供了customer_id，直接使用
        # 否则，如果提供了customer_name，通过名称查找customer_id
        if not customer_id and customer_name:
            tenant = TenantCustomer.query.filter(
                (TenantCustomer.company_name == customer_name) | (TenantCustomer.name == customer_name)
            ).first()
            if tenant:
                customer_id = str(tenant.id)

        if customer_id:
            user = User.query.filter_by(username=username, customer_id=int(customer_id)).first()
        else:
            user = User.query.filter_by(username=username, customer_id=None).first()

        if user and user.check_password(password):
            # 检查试用期
            if user.customer_id:
                tenant = db.session.get(TenantCustomer, user.customer_id)
                if tenant and tenant.trial_expires_at and datetime.utcnow() > tenant.trial_expires_at:
                    flash('试用期已结束，请联系管理员续期', 'warning')
                    return render_template('login.html', tenants=tenants, system_name=system_name)
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            flash('登录成功', 'success')
            if username == 'superadmin':
                return redirect(url_for('tenant_management'))
            return redirect(url_for('index'))
        flash('用户名或密码错误', 'warning')
    return render_template('login.html', tenants=tenants, system_name=system_name)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        company = request.form['company'].strip()
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        if not company or not username or not password:
            flash('请填写所有字段', 'warning')
            return render_template('register.html')
        if TenantCustomer.query.filter_by(name=company).first():
            flash('公司名称已存在', 'warning')
            return render_template('register.html')
        if User.query.filter_by(username=username).first():
            flash(f'用户名 "{username}" 已被使用', 'warning')
            return render_template('register.html')
        tenant = TenantCustomer(
            name=company,
            company_name=company,
            trial_expires_at=datetime.utcnow() + timedelta(days=30)
        )
        db.session.add(tenant)
        db.session.flush()
        admin = User(username=username, role='超级管理员', permissions='all', customer_id=tenant.id)
        admin.set_password(password)
        db.session.add(admin)
        db.session.commit()
        flash('注册成功，试用期30天，请登录', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')


@app.route('/api/user_branding')
def api_user_branding():
    """根据用户名返回该用户所属租户的品牌信息（公司名称+Logo），登录页用"""
    username = request.args.get('username', '').strip()
    if not username:
        return jsonify({'company_name': '', 'logo_url': ''})

    user = User.query.filter_by(username=username).first()
    if not user or not user.customer_id:
        # superadmin 或未关联租户的用户：返回全局配置
        configs = {c.key: c.value for c in SysConfig.query.all()}
        logo_file = configs.get('company_logo_file', '')
        logo_url = f'/static/{logo_file}' if logo_file else ''
        return jsonify({
            'company_name': configs.get('company_name', ''),
            'logo_url': logo_url
        })

    tenant = db.session.get(TenantCustomer, user.customer_id)
    if not tenant:
        return jsonify({'company_name': '', 'logo_url': ''})

    # 租户有自己的品牌信息则用自己的，否则回退到全局配置
    if tenant.company_name or tenant.logo_file:
        logo_url = f'/static/{tenant.logo_file}' if tenant.logo_file else ''
        return jsonify({
            'company_name': tenant.company_name or '',
            'logo_url': logo_url
        })
    else:
        configs = {c.key: c.value for c in SysConfig.query.all()}
        logo_file = configs.get('company_logo_file', '')
        logo_url = f'/static/{logo_file}' if logo_file else ''
        return jsonify({
            'company_name': configs.get('company_name', ''),
            'logo_url': logo_url
        })


@app.route('/logout')
def logout():
    session.clear()
    flash('已退出登录', 'success')
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    # 新增：superadmin 不应访问合同列表，直接重定向到租户管理
    if is_superadmin():
        flash('总超级管理员请在租户管理界面操作，不可查看租户合同数据', 'warning')
        return redirect(url_for('tenant_management'))

    from sqlalchemy import func
    query = Contract.query

    # 新增：数据隔离 - 非superadmin只能看到自己租户的数据
    customer_id = get_current_customer_id()
    if customer_id is not None:
        query = query.filter(Contract.customer_id == customer_id)

    # 新增：数据权限过滤
    user_id = session.get('user_id')
    if user_id and customer_id is not None:
        # 新增：读取"包含已完结合同"选项
        include_finished = request.args.get('include_finished') == '1'
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id, include_finished)
        # None 表示可访问所有合同，不需要过滤
        if accessible_ids is not None:
            if not accessible_ids:
                # 空列表表示没有可访问的合同
                user = db.session.get(User, user_id)
                return render_template('index.html',
                                     contracts=[],
                                     available_years=[],
                                     alerts=[],
                                     stats={
                                         'total_contracts': 0,
                                         'total_price': 0,
                                         'total_paid': 0,
                                         'total_unpaid': 0,
                                         'total_invoiced': 0,
                                         'total_uninvoiced': 0
                                     },
                                     page=1,
                                     per_page=10,
                                     total_pages=0,
                                     user_permissions=get_user_permissions_string(user_id),
                                     is_tenant_user=user.customer_id is not None if user else False)
            else:
                # 过滤可访问的合同
                query = query.filter(Contract.id.in_(accessible_ids))

    # 原有筛选条件
    if request.args.get('project_staff'):
        query = query.filter(Contract.project_staff.like(f"%{request.args.get('project_staff')}%"))
    if request.args.get('customer_name'):
        query = query.filter(Contract.customer_name.like(f"%{request.args.get('customer_name')}%"))

    # 修改：合同类型筛选 - 使用 JOIN 查询 ContractProduct 表
    if request.args.get('contract_type'):
        contract_type = request.args.get('contract_type')
        # 使用 JOIN 确保只查询当前 query 范围内的合同
        query = query.join(ContractProduct, Contract.id == ContractProduct.contract_id).filter(
            ContractProduct.contract_type == contract_type
        ).distinct()

    if request.args.get('status'):
        query = query.filter(Contract.status == request.args.get('status'))
    # 新增：业务类型筛选
    if request.args.get('business_type'):
        query = query.filter(Contract.business_type == request.args.get('business_type'))
    # 新增：签订年份筛选
    if request.args.get('signing_year'):
        query = query.filter(func.strftime('%Y', Contract.signing_date) == request.args.get('signing_year'))

    contracts = query.order_by(Contract.created_at.desc()).all()

    # 发票状态筛选
    if request.args.get('invoice_status'):
        filtered = []
        for contract in contracts:
            has_issued = any(i.invoice_status == '已开具' for i in contract.invoices)
            if request.args.get('invoice_status') == '已开具' and has_issued:
                filtered.append(contract)
            elif request.args.get('invoice_status') == '未开具' and not has_issued:
                filtered.append(contract)
        contracts = filtered

    alerts = []
    today = datetime.now().date()

    for contract in contracts:
        if contract.status == '进行中':
            last_payment = Payment.query.filter_by(contract_id=contract.id).order_by(Payment.payment_date.desc()).first()
            if last_payment:
                days_since = (today - last_payment.payment_date).days
                if days_since > 30:
                    alerts.append(f"{contract.project_name} - 距上次收付款已{days_since}天")

    # 预警筛选
    if request.args.get('alert') == 'yes':
        alert_contracts = [a.split(' - ')[0] for a in alerts]
        contracts = [c for c in contracts if c.project_name in alert_contracts]

    # 新增：获取可用年份列表（用于年份筛选下拉）
    years_raw = db.session.query(func.strftime('%Y', Contract.signing_date)).filter(
        Contract.signing_date.isnot(None)
    ).distinct().order_by(func.strftime('%Y', Contract.signing_date).desc()).all()
    available_years = [int(y[0]) for y in years_raw if y[0]]

    # 筛选结果统计汇总
    stats = {
        'count': len(contracts),
        'total_price': sum(c.total_price for c in contracts),
        'total_paid': sum(c.get_total_paid() for c in contracts),
        'total_unpaid': sum(c.get_unpaid_amount() for c in contracts),
        'total_invoiced': sum(c.get_total_invoiced() for c in contracts),
        'total_uninvoiced': sum(c.get_uninvoiced_amount() for c in contracts),
    }

    # 新增：分页功能
    page = request.args.get('page', 1, type=int)  # 当前页码，默认第1页
    per_page = request.args.get('per_page', 10, type=int)  # 每页显示数量，默认10条

    # 计算总页数
    total_contracts = len(contracts)
    total_pages = (total_contracts + per_page - 1) // per_page  # 向上取整

    # 分页切片
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    contracts_page = contracts[start_idx:end_idx]

    is_tenant_user = get_current_customer_id() is not None
    user_permissions = get_user_permissions_string(session['user_id'])
    return render_template('index.html',
                           contracts=contracts_page,
                           alerts=alerts,
                           available_years=available_years,
                           stats=stats,
                           is_tenant_user=is_tenant_user,
                           user_permissions=user_permissions,
                           page=page,
                           per_page=per_page,
                           total_pages=total_pages,
                           total_contracts=total_contracts)


# ── 新增：合同列表导出 Excel ──
@app.route('/contract/export')
@permission_required('导出EXCEL')
def export_contracts():
    from sqlalchemy import func
    query = Contract.query

    # 新增：租户数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        query = query.filter(Contract.customer_id == customer_id)

    # 数据权限过滤
    user_id = session.get('user_id')
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        # None 表示可访问所有合同，不需要过滤
        if accessible_ids is not None:
            if not accessible_ids:
                # 空列表表示没有可访问的合同，返回空Excel
                flash('没有可导出的合同数据', 'warning')
                return redirect(url_for('index'))
            else:
                # 过滤可访问的合同
                query = query.filter(Contract.id.in_(accessible_ids))

    if request.args.get('project_staff'):
        query = query.filter(Contract.project_staff.like(f"%{request.args.get('project_staff')}%"))
    if request.args.get('customer_name'):
        query = query.filter(Contract.customer_name.like(f"%{request.args.get('customer_name')}%"))

    # 修改：合同类型筛选 - 使用 JOIN 查询 ContractProduct 表
    if request.args.get('contract_type'):
        contract_type = request.args.get('contract_type')
        query = query.join(ContractProduct, Contract.id == ContractProduct.contract_id).filter(
            ContractProduct.contract_type == contract_type
        ).distinct()

    if request.args.get('status'):
        query = query.filter(Contract.status == request.args.get('status'))
    if request.args.get('business_type'):
        query = query.filter(Contract.business_type == request.args.get('business_type'))
    if request.args.get('signing_year'):
        query = query.filter(func.strftime('%Y', Contract.signing_date) == request.args.get('signing_year'))

    contracts = query.order_by(Contract.created_at.desc()).all()

    # 合同级合并列索引（0-based，对应列：合同编号/客户名称/项目名称/合同总价/合同类型/业务类型/项目负责人/销售人员/签订日期/状态/已收付款/未收付款/已开票/未开票）
    MERGE_COLS = [0, 1, 2, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]
    columns = ['合同编号', '客户名称', '项目名称', '产品名称', '型号', '单位', '数量', '单价',
               '合同总价', '发票税率', '合同类型', '业务类型', '项目负责人', '销售人员',
               '签订日期', '状态', '已收付款', '未收付款', '已开票', '未开票']

    # 构建行数据，记录每个合同的起始行和行数（用于合并）
    rows = []
    contract_spans = []  # (start_row, row_count) 1-based，含表头偏移
    for c in contracts:
        products = ContractProduct.query.filter_by(contract_id=c.id).all()
        contract_base = {
            '合同编号': c.contract_number or '',
            '客户名称': c.customer_name,
            '项目名称': c.project_name,
            '合同总价': c.total_price,
            '合同类型': c.contract_type or (products[0].contract_type if products else ''),
            '业务类型': c.business_type or '',
            '项目负责人': c.project_staff or '',
            '销售人员': c.sales_staff or '',
            '签订日期': str(c.signing_date) if c.signing_date else '',
            '状态': c.status or '',
            '已收付款': c.get_total_paid(),
            '未收付款': c.get_unpaid_amount(),
            '已开票': c.get_total_invoiced(),
            '未开票': c.get_uninvoiced_amount(),
        }
        start = len(rows)
        if products:
            for cp in products:
                row = dict(contract_base)
                row.update({'产品名称': cp.product_name or '', '型号': cp.model or '',
                            '单位': cp.unit or '', '数量': cp.quantity,
                            '单价': cp.unit_price, '发票税率': cp.tax_rate})
                rows.append([row.get(col) for col in columns])
        else:
            row = dict(contract_base)
            row.update({'产品名称': c.product_name or '', '型号': c.model or '',
                        '单位': c.unit or '', '数量': c.quantity,
                        '单价': c.unit_price, '发票税率': c.tax_rate})
            rows.append([row.get(col) for col in columns])
        contract_spans.append((start, len(rows) - start))

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = '合同列表'
    ws.append(columns)
    for row in rows:
        ws.append(row)

    # 合并同一合同的合同级列单元格
    for start, span in contract_spans:
        if span > 1:
            data_start = start + 2  # +1 header, +1 1-based
            data_end = start + span + 1
            for col_idx in MERGE_COLS:
                col_letter = get_column_letter(col_idx + 1)
                ws.merge_cells(f'{col_letter}{data_start}:{col_letter}{data_end}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"合同导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/users')
@login_required
def users():
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 新增：superadmin 不应查看租户人员，重定向到租户管理
    if is_superadmin():
        flash('请在租户管理界面管理各租户的用户', 'warning')
        return redirect(url_for('tenant_management'))

    # 新增：客户超级管理员只能看到自己租户下的用户
    customer_id = get_current_customer_id()
    if customer_id is not None:
        users = User.query.filter(User.customer_id == customer_id).all()
    else:
        # superadmin可以看到所有用户
        tenant_id = request.args.get('tenant_id')
        if tenant_id:
            users = User.query.filter(User.customer_id == int(tenant_id)).all()
        else:
            users = User.query.all()

    # 新增：获取租户客户列表（仅superadmin可见）
    tenants = TenantCustomer.query.all() if is_superadmin() else []

    # 新增：获取每个用户的岗位权限信息
    users_position_info = {}
    for user in users:
        position_perms = get_user_position_permissions_info(user.id)
        users_position_info[user.id] = position_perms

    return render_template('users.html', users=users, is_superadmin=is_superadmin(),
                         is_customer_admin=is_customer_admin(), tenants=tenants,
                         users_position_info=users_position_info)


# 新增：租户管理路由（仅superadmin可访问）
@app.route('/tenants')
@login_required
def tenant_management():
    if not is_superadmin():
        flash('权限不足：只有总超级管理员可以管理租户', 'warning')
        return redirect(url_for('index'))
    tenants = TenantCustomer.query.order_by(TenantCustomer.created_at.desc()).all()
    return render_template('tenant_management.html', tenants=tenants)


@app.route('/tenant/<int:tenant_id>/users')
@login_required
def tenant_users(tenant_id):
    """superadmin 查看某租户的用户列表"""
    if not is_superadmin():
        flash('权限不足', 'warning')
        return redirect(url_for('index'))
    tenant = TenantCustomer.query.get_or_404(tenant_id)
    users = User.query.filter_by(customer_id=tenant_id).all()
    return render_template('tenant_users.html', tenant=tenant, users=users)



@app.route('/tenant/create', methods=['POST'])
@login_required
def create_tenant():
    if not is_superadmin():
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    tenant_name = request.form['tenant_name']
    description = request.form.get('description', '')
    admin_username = request.form['admin_username']
    admin_password = request.form['admin_password']
    admin_role = request.form.get('admin_role', '超级管理员')
    selected_perms = request.form.getlist('permissions')
    if selected_perms:
        permissions_str = ','.join(selected_perms)
    else:
        permissions_str = 'all'

    # 检查租户名称是否重复
    if TenantCustomer.query.filter_by(name=tenant_name).first():
        flash('租户名称已存在', 'warning')
        return redirect(url_for('tenant_management'))

    # 检查管理员账号是否重复
    existing_user = User.query.filter_by(username=admin_username).first()
    if existing_user:
        # 如果该账号关联的租户已不存在（孤儿账号），自动清除并允许继续
        if existing_user.customer_id is not None and db.session.get(TenantCustomer, existing_user.customer_id) is None:
            db.session.delete(existing_user)
            db.session.flush()
        else:
            flash(f'管理员账号"{admin_username}"已存在，请换一个账号名称', 'warning')
            return redirect(url_for('tenant_management'))

    # 创建租户
    tenant = TenantCustomer(name=tenant_name, description=description)
    db.session.add(tenant)
    db.session.flush()

    # 创建该租户的管理员（角色和权限由表单指定）
    admin = User(
        username=admin_username,
        role=admin_role,
        permissions=permissions_str,
        customer_id=tenant.id
    )
    admin.set_password(admin_password)
    db.session.add(admin)
    db.session.commit()

    flash(f'租户"{tenant_name}"创建成功，管理员账号：{admin_username}', 'success')
    return redirect(url_for('tenant_management'))


# 新增：设置租户品牌信息（公司名称+Logo）
@app.route('/tenant/<int:tenant_id>/branding', methods=['POST'])
@login_required
def tenant_branding(tenant_id):
    if not is_superadmin():
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    tenant = TenantCustomer.query.get_or_404(tenant_id)
    tenant.company_name = request.form.get('company_name', '').strip()

    # 处理 Logo 文件上传
    if 'logo_file' in request.files:
        f = request.files['logo_file']
        if f.filename:
            ext = os.path.splitext(f.filename)[1]
            logo_filename = f'tenant_{tenant_id}_logo{ext}'
            f.save(os.path.join('static', logo_filename))
            tenant.logo_file = logo_filename

    db.session.commit()
    flash(f'租户"{tenant.name}"的品牌信息已更新', 'success')
    return redirect(url_for('tenant_management'))


@app.route('/user/<int:id>/reset_password', methods=['POST'])
@login_required
def reset_user_password(id):
    """superadmin 重置租户用户密码"""
    if not is_superadmin():
        flash('权限不足', 'warning')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    user.set_password(request.form['password'])
    db.session.commit()
    flash(f'用户"{user.username}"密码已重置', 'success')
    return redirect(url_for('tenant_users', tenant_id=user.customer_id))



@app.route('/tenant/<int:tenant_id>/edit', methods=['POST'])
@login_required
def tenant_edit(tenant_id):
    if not is_superadmin():
        flash('权限不足', 'warning')
        return redirect(url_for('index'))
    tenant = TenantCustomer.query.get_or_404(tenant_id)
    tenant.name = request.form.get('tenant_name', tenant.name).strip()
    tenant.description = request.form.get('description', tenant.description)

    # 新增：支持修改管理员账号和密码
    new_username = request.form.get('admin_username', '').strip()
    new_password = request.form.get('admin_password', '').strip()
    if new_username or new_password:
        admin = User.query.filter_by(customer_id=tenant_id).first()
        if admin:
            if new_username and new_username != admin.username:
                if User.query.filter_by(username=new_username).first():
                    flash(f'账号"{new_username}"已存在', 'warning')
                    return redirect(url_for('tenant_management'))
                admin.username = new_username
            if new_password:
                admin.set_password(new_password)

    db.session.commit()
    flash('租户信息已更新', 'success')
    return redirect(url_for('tenant_management'))


@app.route('/tenant/<int:tenant_id>/delete', methods=['POST'])
@login_required
def tenant_delete(tenant_id):
    if not is_superadmin():
        flash('权限不足', 'warning')
        return redirect(url_for('index'))
    tenant = TenantCustomer.query.get_or_404(tenant_id)
    name = tenant.name
    User.query.filter_by(customer_id=tenant_id).delete(synchronize_session='fetch')
    db.session.flush()
    db.session.delete(tenant)
    db.session.commit()
    flash(f'租户"{name}"已删除', 'success')
    return redirect(url_for('tenant_management'))


# ========== 组织结构管理 ==========

@app.route('/organizations')
@login_required
def organizations():
    """组织结构列表（客户超级管理员可访问）"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    if customer_id is None:
        flash('superadmin 无需管理组织结构', 'warning')
        return redirect(url_for('index'))

    # 获取当前租户的所有组织（树形结构）
    orgs = Organization.query.filter_by(customer_id=customer_id).order_by(Organization.created_at).all()

    # 获取当前租户的所有用户
    users = User.query.filter_by(customer_id=customer_id).all()

    return render_template('organizations.html', organizations=orgs, users=users)


@app.route('/organization/create', methods=['POST'])
@login_required
def create_organization():
    """创建组织"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    if customer_id is None:
        flash('superadmin 无需创建组织', 'warning')
        return redirect(url_for('index'))

    name = request.form['name']
    description = request.form.get('description', '')
    parent_id = request.form.get('parent_id')
    is_virtual = request.form.get('is_virtual') == 'on'  # 虚拟组织复选框

    if parent_id and parent_id.strip():
        parent_id = int(parent_id)
    else:
        parent_id = None

    selected_perms = request.form.getlist('permissions')
    org = Organization(
        name=name,
        description=description,
        parent_id=parent_id,
        customer_id=customer_id,
        permissions=','.join(selected_perms) if selected_perms else None,
        is_virtual=is_virtual
    )
    db.session.add(org)
    db.session.commit()

    org_type = '虚拟组织' if is_virtual else '组织'
    flash(f'{org_type}"{name}"创建成功', 'success')
    return redirect(url_for('organizations'))


@app.route('/organization/<int:org_id>/edit', methods=['POST'])
@login_required
def edit_organization(org_id):
    """编辑组织"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    org = Organization.query.get_or_404(org_id)

    # 验证权限：只能编辑自己租户的组织
    if org.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('organizations'))

    org.name = request.form['name']
    org.description = request.form.get('description', '')
    parent_id = request.form.get('parent_id')
    org.is_virtual = request.form.get('is_virtual') == 'on'  # 虚拟组织复选框

    if parent_id and parent_id.strip():
        org.parent_id = int(parent_id)
    else:
        org.parent_id = None

    selected_perms = request.form.getlist('permissions')
    org.permissions = ','.join(selected_perms) if selected_perms else None

    db.session.commit()
    flash(f'组织"{org.name}"已更新', 'success')
    return redirect(url_for('organizations'))


@app.route('/organization/<int:org_id>/delete', methods=['POST'])
@login_required
def delete_organization(org_id):
    """删除组织"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    org = Organization.query.get_or_404(org_id)

    # 验证权限
    if org.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('organizations'))

    # 检查是否有子组织
    if org.children:
        flash('该组织下有子组织，无法删除', 'warning')
        return redirect(url_for('organizations'))

    # 检查是否有成员
    if org.members:
        flash('该组织下有成员，无法删除', 'warning')
        return redirect(url_for('organizations'))

    db.session.delete(org)
    db.session.commit()
    flash('组织已删除', 'success')
    return redirect(url_for('organizations'))


@app.route('/organization/transfer', methods=['POST'])
@login_required
def transfer_user():
    """人员调动（调入/调出组织）"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    user_id = int(request.form['user_id'])
    target_org_id = request.form.get('target_org_id')

    user = User.query.get_or_404(user_id)

    # 验证权限：只能调动自己租户的用户
    if user.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('organizations'))

    if target_org_id and target_org_id.strip():
        target_org_id = int(target_org_id)
        # 验证目标组织属于当前租户
        target_org = db.session.get(Organization, target_org_id)
        if not target_org or target_org.customer_id != customer_id:
            flash('目标组织不存在或权限不足', 'warning')
            return redirect(url_for('organizations'))
        user.organization_id = target_org_id

        # 新增：同步更新 UserOrganization 表的主组织标记
        # 先取消所有主组织标记
        UserOrganization.query.filter_by(user_id=user.id, is_primary=True).update({'is_primary': False})

        # 查找或创建目标组织的记录
        existing_user_org = UserOrganization.query.filter_by(
            user_id=user.id,
            organization_id=target_org_id
        ).first()

        if existing_user_org:
            # 如果记录已存在，更新为主组织
            existing_user_org.is_primary = True
        else:
            # 如果记录不存在，创建新记录
            user_org = UserOrganization(
                user_id=user.id,
                organization_id=target_org_id,
                is_primary=True
            )
            db.session.add(user_org)

        # 新增：同步更新用户岗位的组织归属
        # 将用户的所有岗位都更新到新组织
        user_positions = UserPosition.query.filter_by(user_id=user.id).all()
        for up in user_positions:
            up.organization_id = target_org_id

        # 合并组织权限（保留用户原有权限）
        if target_org.permissions:
            # 获取用户当前权限集合
            current_perms = set()
            if user.permissions and user.permissions != 'all':
                current_perms = set(user.permissions.split(','))

            # 获取组织权限集合
            org_perms = set(target_org.permissions.split(','))

            # 合并权限（用户权限 + 组织权限）
            merged_perms = current_perms.union(org_perms)
            user.permissions = ','.join(sorted(merged_perms))
        flash(f'用户"{user.username}"已调入组织"{target_org.name}"', 'success')
    else:
        # 调出组织（设为 None）
        user.organization_id = None
        # 新增：同时清除 UserOrganization 表中的主组织标记
        UserOrganization.query.filter_by(user_id=user.id, is_primary=True).update({'is_primary': False})
        flash(f'用户"{user.username}"已调出组织', 'success')

    db.session.commit()
    return redirect(url_for('organizations'))


@app.route('/user/new', methods=['GET', 'POST'])
@login_required
def new_user():
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))
    if request.method == 'POST':
        role = request.form['role']
        # 只有 superadmin 可以创建超级管理员角色用户
        if role == '超级管理员' and not is_superadmin():
            flash('权限不足：只有 superadmin 可以创建超级管理员账户', 'warning')
            return redirect(url_for('users'))

        # 新增：客户超级管理员创建的用户自动继承其customer_id
        customer_id = get_current_customer_id()
        if customer_id is not None and role == '超级管理员':
            flash('客户超级管理员不能创建超级管理员账户', 'warning')
            return redirect(url_for('users'))

        # superadmin创建客户超级管理员时需要指定租户
        if is_superadmin() and role == '超级管理员' and request.form.get('customer_id'):
            customer_id = int(request.form['customer_id'])

        username = request.form['username']
        existing = User.query.filter_by(username=username, customer_id=customer_id).first()
        if existing:
            flash(f'用户名 "{username}" 已存在，请使用其他用户名', 'danger')
            tenants = TenantCustomer.query.all() if is_superadmin() else []
            return render_template('user_form.html', is_superadmin=is_superadmin(),
                                 is_customer_admin=is_customer_admin(), tenants=tenants)
        user = User(
            username=username,
            role=role,
            permissions=','.join(request.form.getlist('permissions')),
            customer_id=customer_id
        )
        user.set_password(request.form['password'])
        db.session.add(user)
        db.session.commit()
        flash('用户创建成功', 'success')
        return redirect(url_for('users'))

    tenants = TenantCustomer.query.all() if is_superadmin() else []
    return render_template('user_form.html', is_superadmin=is_superadmin(),
                         is_customer_admin=is_customer_admin(), tenants=tenants)


@app.route('/user/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(id):
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    # 非 superadmin 不能编辑超级管理员账户
    if user.role == '超级管理员' and not is_superadmin():
        flash('权限不足：只有 superadmin 可以修改超级管理员账户', 'warning')
        return redirect(url_for('users'))
    # superadmin 自身不可修改角色
    if user.username == 'superadmin' and not is_superadmin():
        flash('权限不足', 'warning')
        return redirect(url_for('users'))
    if request.method == 'POST':
        new_role = request.form['role']
        if new_role == '超级管理员' and not is_superadmin():
            flash('权限不足：只有 superadmin 可以设置超级管理员角色', 'warning')
            return redirect(url_for('users'))
        user.username = request.form['username']
        user.role = new_role
        user.permissions = ','.join(request.form.getlist('permissions'))
        if request.form.get('password'):
            user.set_password(request.form['password'])
        db.session.commit()
        flash('用户更新成功', 'success')
        return redirect(url_for('users'))
    return render_template('user_form.html', user=user, is_superadmin=is_superadmin())


@app.route('/user/<int:id>/delete', methods=['POST'])
@login_required
def delete_user(id):
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    if user.username in ('admin', 'superadmin'):
        flash('不能删除系统内置管理员账户', 'warning')
        return redirect(url_for('users'))
    # 非 superadmin 不能删除超级管理员用户
    if user.role == '超级管理员' and not is_superadmin():
        flash('权限不足：只有 superadmin 可以删除超级管理员账户', 'warning')
        return redirect(url_for('users'))
    db.session.delete(user)
    db.session.commit()
    flash('用户删除成功', 'success')
    return redirect(url_for('users'))


# ==================== 岗位管理 ====================

@app.route('/positions')
@login_required
def positions():
    """岗位管理列表（客户超级管理员可访问）"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    if customer_id is None:
        flash('superadmin 无需管理岗位', 'warning')
        return redirect(url_for('index'))

    # 获取当前租户的所有岗位
    positions = Position.query.filter_by(customer_id=customer_id).order_by(Position.created_at).all()

    # 获取当前租户的所有合同（用于岗位赋权）
    contracts = Contract.query.filter_by(customer_id=customer_id).all()

    # 获取当前租户的所有组织
    organizations = Organization.query.filter_by(customer_id=customer_id).all()

    # 获取当前租户的所有用户
    users = User.query.filter_by(customer_id=customer_id).all()

    # 新增：构建用户岗位映射表
    user_positions_map = {}
    for user in users:
        user_positions = UserPosition.query.filter_by(user_id=user.id).all()
        positions_list = []
        for up in user_positions:
            pos = db.session.get(Position, up.position_id)
            org = db.session.get(Organization, up.organization_id) if up.organization_id else None
            if pos:
                positions_list.append({
                    'position_name': pos.name,
                    'organization_name': org.name if org else None,
                    'is_primary': up.is_primary
                })
        user_positions_map[user.id] = positions_list

    # 新增：统计每个岗位的人员数
    user_positions_count = {}
    for position in positions:
        count = UserPosition.query.filter_by(position_id=position.id).count()
        user_positions_count[position.id] = count

    return render_template('positions.html', positions=positions, contracts=contracts,
                         organizations=organizations, users=users, user_positions_map=user_positions_map,
                         user_positions_count=user_positions_count)


@app.route('/position/create', methods=['POST'])
@login_required
def create_position():
    """创建岗位"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    if customer_id is None:
        flash('superadmin 无需创建岗位', 'warning')
        return redirect(url_for('index'))

    name = request.form['name']
    description = request.form.get('description', '')
    data_scope = request.form.get('data_scope', 'all')

    selected_perms = request.form.getlist('function_permissions')

    position = Position(
        name=name,
        description=description,
        customer_id=customer_id,
        function_permissions=','.join(selected_perms) if selected_perms else None,
        data_scope=data_scope
    )
    db.session.add(position)
    db.session.commit()

    flash(f'岗位"{name}"创建成功', 'success')
    return redirect(url_for('positions'))


@app.route('/position/<int:position_id>/edit', methods=['POST'])
@login_required
def edit_position(position_id):
    """编辑岗位"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    position = Position.query.get_or_404(position_id)

    # 验证权限：只能编辑自己租户的岗位
    if position.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('positions'))

    position.name = request.form['name']
    position.description = request.form.get('description', '')
    position.data_scope = request.form.get('data_scope', 'all')

    selected_perms = request.form.getlist('function_permissions')
    new_function_permissions = ','.join(selected_perms) if selected_perms else None

    position.function_permissions = new_function_permissions

    # 新增：岗位优先模型 —— 修改岗位权限后，清空该岗位所有用户的自身权限，
    # 使这些用户的功能权限完全由岗位决定，确保岗位权限修改立即对用户生效。
    # 超级管理员始终拥有全部权限，不受影响。
    user_positions = UserPosition.query.filter_by(position_id=position_id).all()
    affected_users = []
    for up in user_positions:
        user = db.session.get(User, up.user_id)
        if user and user.role != '超级管理员' and user.permissions:
            user.permissions = None
            affected_users.append(user.username)

    db.session.commit()

    if affected_users:
        preview = '、'.join(affected_users[:3]) + ('等' if len(affected_users) > 3 else '')
        flash(f'岗位"{position.name}"已更新。该岗位 {len(affected_users)} 名用户的权限已改为完全由岗位控制并立即生效（{preview}）。如用户当前页面未更新，请刷新页面。', 'success')
    else:
        flash(f'岗位"{position.name}"已更新，该岗位用户的权限已同步生效。如用户当前页面权限未更新，请刷新页面。', 'success')
    return redirect(url_for('positions'))


@app.route('/position/<int:position_id>/delete', methods=['POST'])
@login_required
def delete_position(position_id):
    """删除岗位"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    position = Position.query.get_or_404(position_id)

    # 验证权限：只能删除自己租户的岗位
    if position.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('positions'))

    # 删除关联的用户岗位和权限
    UserPosition.query.filter_by(position_id=position_id).delete()
    PositionContractPermission.query.filter_by(position_id=position_id).delete()

    db.session.delete(position)
    db.session.commit()
    flash(f'岗位"{position.name}"已删除', 'success')
    return redirect(url_for('positions'))


@app.route('/position/<int:position_id>/assign_user', methods=['POST'])
@login_required
def assign_user_to_position(position_id):
    """为用户分配岗位"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    position = Position.query.get_or_404(position_id)

    if position.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('positions'))

    user_id = int(request.form['user_id'])
    org_id = request.form.get('organization_id')
    org_id = int(org_id) if org_id and org_id.strip() else None
    is_primary = request.form.get('is_primary') == 'on'

    # 新增：如果岗位的数据权限是"本组织合同"，但未指定组织，给出提示
    if position.data_scope == 'org' and not org_id:
        flash(f'岗位"{position.name}"的数据权限为"本组织合同"，必须指定所属组织，否则用户无法访问组织数据', 'warning')
        return redirect(url_for('positions'))

    # 检查是否已存在
    existing = UserPosition.query.filter_by(user_id=user_id, position_id=position_id, organization_id=org_id).first()
    if existing:
        flash('该用户已拥有此岗位', 'warning')
        return redirect(url_for('positions'))

    # 如果设置为主岗位，清除该用户的其他主岗位
    if is_primary:
        UserPosition.query.filter_by(user_id=user_id, is_primary=True).update({'is_primary': False})

    user_position = UserPosition(
        user_id=user_id,
        position_id=position_id,
        organization_id=org_id,
        is_primary=is_primary
    )
    db.session.add(user_position)

    # 新增：同步更新 User.organization_id 和 UserOrganization 表
    user = db.session.get(User, user_id)
    if org_id:
        # 如果指定了组织
        if is_primary or not user.organization_id:
            # 如果是主岗位，或者用户还没有主组织，则更新
            user.organization_id = org_id

        # 确保 UserOrganization 表中有记录
        user_org = UserOrganization.query.filter_by(
            user_id=user_id,
            organization_id=org_id
        ).first()

        if not user_org:
            # 如果是主岗位，先取消其他主组织
            if is_primary:
                UserOrganization.query.filter_by(
                    user_id=user_id,
                    is_primary=True
                ).update({'is_primary': False})

            user_org = UserOrganization(
                user_id=user_id,
                organization_id=org_id,
                is_primary=is_primary
            )
            db.session.add(user_org)
        elif is_primary:
            # 更新为主组织
            UserOrganization.query.filter_by(
                user_id=user_id,
                is_primary=True
            ).update({'is_primary': False})
            user_org.is_primary = True

    db.session.commit()

    # 新增：分配岗位后，检查用户是否能访问合同数据，给出智能提示
    user_name = user.username
    position_name = position.name

    # 检查用户能看到的合同
    accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)

    if accessible_ids is None:
        # None = 全部合同，权限正常
        flash(f'已为用户"{user_name}"分配岗位"{position_name}"，该用户可访问全部合同。', 'success')
    elif len(accessible_ids) == 0:
        # 空列表 = 看不到任何合同，给出详细提示
        if position.data_scope == 'org' and org_id:
            org = db.session.get(Organization, org_id)
            org_name = org.name if org else f'组织{org_id}'
            # 检查该组织是否有合同
            org_contract_count = Contract.query.filter_by(customer_id=customer_id, organization_id=org_id).count()
            if org_contract_count == 0:
                flash(f'⚠️ 已为用户"{user_name}"分配岗位"{position_name}"（数据范围：本组织）。但组织"{org_name}"下目前没有合同，该用户暂时无法看到任何合同。建议：前往"合同组织分配"页面，将合同分配给"{org_name}"组织。', 'warning')
            else:
                flash(f'⚠️ 已为用户"{user_name}"分配岗位"{position_name}"。组织"{org_name}"有{org_contract_count}条合同，但该用户当前无法访问（可能合同都已完结，或用户未被显式授权）。', 'warning')
        else:
            flash(f'⚠️ 已为用户"{user_name}"分配岗位"{position_name}"，但该用户当前无法访问任何合同。建议：检查岗位的数据范围设置，或为该岗位授权特定合同。', 'warning')
    else:
        # 能看到部分合同
        flash(f'已为用户"{user_name}"分配岗位"{position_name}"，该用户可访问 {len(accessible_ids)} 条合同。', 'success')

    return redirect(url_for('positions'))


@app.route('/position/<int:position_id>/grant_contract', methods=['POST'])
@login_required
def grant_contract_to_position(position_id):
    """为岗位授权访问特定合同"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    position = Position.query.get_or_404(position_id)

    if position.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('positions'))

    contract_id = int(request.form['contract_id'])
    permission_type = request.form.get('permission_type', 'view')

    # 检查是否已存在
    existing = PositionContractPermission.query.filter_by(
        position_id=position_id,
        contract_id=contract_id
    ).first()
    if existing:
        # 更新权限类型
        existing.permission_type = permission_type
        db.session.commit()
        flash('合同权限已更新', 'success')
    else:
        pcp = PositionContractPermission(
            position_id=position_id,
            contract_id=contract_id,
            permission_type=permission_type
        )
        db.session.add(pcp)
        db.session.commit()
        flash('合同权限已授予', 'success')

    return redirect(url_for('positions'))


@app.route('/user_position/<int:up_id>/remove', methods=['POST'])
@login_required
def remove_user_position(up_id):
    """移除用户的岗位"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    up = UserPosition.query.get_or_404(up_id)
    db.session.delete(up)
    db.session.commit()
    flash('岗位已移除', 'success')
    return redirect(url_for('positions'))


# 新增：API - 获取岗位下的用户列表
@app.route('/api/position/<int:position_id>/users')
@login_required
def api_position_users(position_id):
    """获取指定岗位下的用户列表（用于分配用户弹窗）"""
    if session.get('role') != '超级管理员':
        return jsonify([])

    customer_id = get_current_customer_id()
    position = Position.query.get_or_404(position_id)

    # 验证权限
    if position.customer_id != customer_id:
        return jsonify([])

    # 查询该岗位下的所有用户
    user_positions = UserPosition.query.filter_by(position_id=position_id).all()
    result = []
    for up in user_positions:
        user = db.session.get(User, up.user_id)
        org = db.session.get(Organization, up.organization_id) if up.organization_id else None
        if user:
            result.append({
                'user_position_id': up.id,
                'user_id': user.id,
                'username': user.username,
                'role': user.role,
                'organization_name': org.name if org else None,
                'is_primary': up.is_primary
            })

    return jsonify(result)


@app.route('/api/position/<int:position_id>')
@login_required
def api_position_detail(position_id):
    """新增：获取岗位详情（用于编辑弹窗实时回显当前功能权限，确保显示与实际一致）"""
    if session.get('role') != '超级管理员':
        return jsonify({})

    customer_id = get_current_customer_id()
    position = Position.query.get_or_404(position_id)
    if position.customer_id != customer_id:
        return jsonify({})

    return jsonify({
        'id': position.id,
        'name': position.name,
        'description': position.description or '',
        'function_permissions': position.function_permissions or '',
        'data_scope': position.data_scope or 'all'
    })


# ==================== 虚拟组织管理 ====================


# ==================== 合同组织分配管理 ====================

@app.route('/contract_organization_assignment')
@login_required
def contract_organization_assignment():
    """合同组织分配管理界面"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()

    # 统计各组织的合同数量
    from sqlalchemy import func
    organization_stats = db.session.query(
        Contract.organization_id,
        func.count(Contract.id)
    ).filter(Contract.customer_id == customer_id).group_by(Contract.organization_id).all()

    # 获取所有组织
    organizations = Organization.query.filter_by(customer_id=customer_id).all()
    organizations_dict = {org.id: org.name for org in organizations}

    # 获取未分配组织的合同
    unassigned_contracts = Contract.query.filter_by(
        customer_id=customer_id,
        organization_id=None
    ).all()

    # 为未分配合同添加创建人名称
    for contract in unassigned_contracts:
        if contract.created_by:
            creator = db.session.get(User, contract.created_by)
            contract.creator_name = creator.username if creator else None
        else:
            contract.creator_name = None

    # 合同总数
    total_contracts = Contract.query.filter_by(customer_id=customer_id).count()

    return render_template('contract_organization_assignment.html',
                          organization_stats=organization_stats,
                          organizations_dict=organizations_dict,
                          organizations=organizations,
                          unassigned_contracts=unassigned_contracts,
                          total_contracts=total_contracts)


@app.route('/assign_all_contracts_to_org', methods=['POST'])
@login_required
def assign_all_contracts_to_org():
    """将所有合同分配到同一个组织"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    org_id = int(request.form['organization_id'])

    # 验证组织属于当前租户
    org = Organization.query.get_or_404(org_id)
    if org.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('contract_organization_assignment'))

    # 更新所有合同
    result = Contract.query.filter_by(customer_id=customer_id).update({'organization_id': org_id})
    db.session.commit()

    flash(f'已将 {result} 个合同分配到"{org.name}"', 'success')
    return redirect(url_for('contract_organization_assignment'))


@app.route('/auto_assign_by_staff', methods=['POST'])
@login_required
def auto_assign_by_staff():
    """根据项目负责人自动分配"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    contracts = Contract.query.filter_by(customer_id=customer_id).all()

    assigned_count = 0
    for contract in contracts:
        if contract.project_staff:
            # 取第一个项目负责人
            staff_name = contract.project_staff.split(',')[0].strip()
            user = User.query.filter_by(username=staff_name, customer_id=customer_id).first()

            if user and user.organization_id:
                contract.organization_id = user.organization_id
                assigned_count += 1

    db.session.commit()
    flash(f'已根据项目负责人自动分配 {assigned_count} 个合同', 'success')
    return redirect(url_for('contract_organization_assignment'))


@app.route('/auto_assign_by_creator', methods=['POST'])
@login_required
def auto_assign_by_creator():
    """根据创建人自动分配"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    contracts = Contract.query.filter_by(customer_id=customer_id).all()

    assigned_count = 0
    for contract in contracts:
        if contract.created_by:
            user = db.session.get(User, contract.created_by)
            if user and user.organization_id:
                contract.organization_id = user.organization_id
                assigned_count += 1

    db.session.commit()
    flash(f'已根据创建人自动分配 {assigned_count} 个合同', 'success')
    return redirect(url_for('contract_organization_assignment'))


@app.route('/clear_all_assignments', methods=['POST'])
@login_required
def clear_all_assignments():
    """清除所有合同的组织分配"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    result = Contract.query.filter_by(customer_id=customer_id).update({'organization_id': None})
    db.session.commit()

    flash(f'已清除 {result} 个合同的组织分配', 'warning')
    return redirect(url_for('contract_organization_assignment'))


@app.route('/assign_single_contract', methods=['POST'])
@login_required
def assign_single_contract():
    """单个合同分配组织"""
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()
    contract_id = int(request.form['contract_id'])
    org_id = int(request.form['organization_id'])

    contract = Contract.query.get_or_404(contract_id)
    org = Organization.query.get_or_404(org_id)

    # 验证权限
    if contract.customer_id != customer_id or org.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('contract_organization_assignment'))

    contract.organization_id = org_id
    db.session.commit()

    flash(f'合同"{contract.contract_number}"已分配到"{org.name}"', 'success')
    return redirect(url_for('contract_organization_assignment'))


@app.route('/api/contracts_by_organization/<org_id>')
@login_required
def api_contracts_by_organization(org_id):
    """获取指定组织的合同列表（API）"""
    if session.get('role') != '超级管理员':
        return jsonify([])

    customer_id = get_current_customer_id()

    # 查询条件
    if org_id == 'null':
        # 未分配组织的合同
        contracts = Contract.query.filter_by(
            customer_id=customer_id,
            organization_id=None
        ).all()
    else:
        # 指定组织的合同
        org_id = int(org_id)
        contracts = Contract.query.filter_by(
            customer_id=customer_id,
            organization_id=org_id
        ).all()

    # 转换为JSON
    result = []
    for c in contracts:
        result.append({
            'id': c.id,
            'contract_number': c.contract_number,
            'project_name': c.project_name,
            'customer_name': c.customer_name,
            'project_staff': c.project_staff,
            'total_price': float(c.total_price) if c.total_price else None,
            'signing_date': c.signing_date.strftime('%Y-%m-%d') if c.signing_date else None
        })

    return jsonify(result)


# ==================== 原有虚拟组织管理 ====================


# 新增：带业务类型的新建客户路由
@app.route('/customer/new/<business_type>', methods=['GET', 'POST'])
@login_required
@permission_required('增加')
def new_customer_with_type(business_type):
    """新建客户（指定业务类型：采购/销售）"""
    if business_type not in ['采购', '销售']:
        flash('业务类型错误', 'warning')
        return redirect(url_for('customers'))

    if request.method == 'POST':
        customer = Customer(
            name=request.form['name'],
            province=request.form.get('province'),
            region=request.form.get('region'),
            credit_code=request.form.get('credit_code'),
            business_type=business_type,
            customer_id=get_current_customer_id()
        )
        db.session.add(customer)
        db.session.commit()
        flash(f'{business_type}客户添加成功', 'success')
        return redirect(url_for('customer_archive', business_type=business_type))

    return render_template('customer_form.html', business_type=business_type)


# 新增：客户档案路由（按业务类型筛选）
@app.route('/customer/archive/<business_type>')
@login_required
def customer_archive(business_type):
    """客户档案（采购/销售）"""
    if business_type not in ['采购', '销售']:
        flash('业务类型错误', 'warning')
        return redirect(url_for('customers'))

    query = Customer.query.filter(Customer.business_type == business_type)

    # 数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        query = query.filter(Customer.customer_id == customer_id)

    # 筛选条件
    if request.args.get('name'):
        query = query.filter(Customer.name.like(f"%{request.args.get('name')}%"))
    if request.args.get('province'):
        query = query.filter(Customer.province.like(f"%{request.args.get('province')}%"))

    customers_list = query.order_by(Customer.province, Customer.name).all()

    # 获取用户权限
    user_permissions = get_user_permissions_string(session['user_id'])

    return render_template('customer_archive.html',
                         customers=customers_list,
                         business_type=business_type,
                         user_permissions=user_permissions)


@app.route('/customers')
@login_required
def customers():
    user_id = session.get('user_id')

    # 检查是否有客户相关权限
    perms = get_user_function_permissions(user_id)
    has_customer_perm = any('客户' in p for p in perms)

    if not has_customer_perm and perms != 'all':
        flash('您没有客户管理权限，请联系管理员', 'warning')
        return redirect(url_for('index'))

    query = Customer.query

    # 新增：数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        query = query.filter(Customer.customer_id == customer_id)

    if request.args.get('name'):
        query = query.filter(Customer.name.like(f"%{request.args.get('name')}%"))
    if request.args.get('province'):
        query = query.filter(Customer.province.like(f"%{request.args.get('province')}%"))
    customers = query.order_by(Customer.province, Customer.name).all()
    return render_template('customers.html', customers=customers)


@app.route('/customer/new', methods=['POST'])
@login_required
@permission_required('增加')
def new_customer():
    customer = Customer(
        name=request.form['name'],
        province=request.form.get('province'),
        region=request.form.get('region'),
        credit_code=request.form.get('credit_code'),
        customer_id=get_current_customer_id()  # 新增：自动关联租户
    )
    db.session.add(customer)
    db.session.commit()
    flash('客户添加成功', 'success')
    return redirect(url_for('customers'))


@app.route('/customer/<int:id>/edit', methods=['POST'])
@login_required
@permission_required('修改')
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    customer.name = request.form['name']
    customer.province = request.form.get('province')
    customer.region = request.form.get('region')
    customer.credit_code = request.form.get('credit_code')
    db.session.commit()
    flash('客户更新成功', 'success')
    return redirect(url_for('customers'))


@app.route('/customer/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('删除')
def delete_customer(id):
    customer = Customer.query.get_or_404(id)
    db.session.delete(customer)
    db.session.commit()
    flash('客户删除成功', 'success')
    return redirect(url_for('customers'))


@app.route('/api/customers/search')
@login_required
def search_customers():
    query = request.args.get('q', '')
    q = Customer.query.filter(Customer.name.like(f'%{query}%'))

    # 新增：数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        q = q.filter(Customer.customer_id == customer_id)

    customers = q.limit(10).all()
    return jsonify([{'id': c.id, 'name': c.name, 'province': c.province} for c in customers])


# ── 新增：项目负责人关键字搜索（从合同记录中提取，支持逗号分隔的多人） ──
@app.route('/api/project_staff/search')
@login_required
def search_project_staff():
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify([])

    # 新增：租户数据隔离
    q = db.session.query(Contract.project_staff)
    customer_id = get_current_customer_id()
    if customer_id is not None:
        q = q.filter(Contract.customer_id == customer_id)

    rows = q.filter(
        Contract.project_staff.isnot(None),
        Contract.project_staff != '',
        Contract.project_staff.like(f'%{query}%')
    ).all()
    names = set()
    for (staff_str,) in rows:
        if staff_str:
            for name in staff_str.split(','):
                name = name.strip()
                if name and query.lower() in name.lower():
                    names.add(name)
    return jsonify(sorted(list(names))[:10])


# 新增：带产品类型的新建产品路由
@app.route('/product/new/<product_type>', methods=['GET', 'POST'])
@login_required
@permission_required('增加')
def new_product_with_type(product_type):
    """新建产品（指定产品类型：硬件设备/软件/技术服务/技术开发）"""
    valid_types = ['硬件设备', '软件', '技术服务', '技术开发']
    if product_type not in valid_types:
        flash('产品类型错误', 'warning')
        return redirect(url_for('products'))

    if request.method == 'POST':
        product = Product(
            name=request.form['name'],
            category=product_type,  # 使用URL参数指定的产品类型
            model=request.form.get('model'),
            unit=request.form.get('unit'),
            tax_rate=float(request.form['tax_rate']) if request.form.get('tax_rate') else None,
            ref_quantity=float(request.form['ref_quantity']) if request.form.get('ref_quantity') else None,
            ref_unit_price=float(request.form['ref_unit_price']) if request.form.get('ref_unit_price') else None,
            customer_id=get_current_customer_id()
        )
        db.session.add(product)
        db.session.commit()
        flash(f'{product_type}产品添加成功', 'success')
        return redirect(url_for('product_archive', product_type=product_type))

    return render_template('product_form.html', product_type=product_type)


# 新增：产品档案路由（按产品类型筛选）
@app.route('/product/archive/<product_type>')
@login_required
def product_archive(product_type):
    """产品档案（硬件设备/软件/技术服务/技术开发）"""
    valid_types = ['硬件设备', '软件', '技术服务', '技术开发']
    if product_type not in valid_types:
        flash('产品类型错误', 'warning')
        return redirect(url_for('products'))

    query = Product.query.filter(Product.category == product_type)

    # 数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        query = query.filter(Product.customer_id == customer_id)

    # 筛选条件
    if request.args.get('name'):
        query = query.filter(Product.name.like(f"%{request.args.get('name')}%"))
    if request.args.get('model'):
        query = query.filter(Product.model.like(f"%{request.args.get('model')}%"))

    products_list = query.order_by(Product.name).all()

    # 获取用户权限
    user_permissions = get_user_permissions_string(session['user_id'])

    return render_template('product_archive.html',
                         products=products_list,
                         product_type=product_type,
                         user_permissions=user_permissions)


@app.route('/products')
@login_required
def products():
    user_id = session.get('user_id')

    # 检查是否有产品相关权限
    perms = get_user_function_permissions(user_id)
    has_product_perm = any('产品' in p for p in perms)

    if not has_product_perm and perms != 'all':
        flash('您没有产品管理权限，请联系管理员', 'warning')
        return redirect(url_for('index'))

    query = Product.query

    # 新增：数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        query = query.filter(Product.customer_id == customer_id)

    if request.args.get('name'):
        query = query.filter(Product.name.like(f"%{request.args.get('name')}%"))
    if request.args.get('category'):
        query = query.filter(Product.category == request.args.get('category'))
    if request.args.get('model'):
        query = query.filter(Product.model.like(f"%{request.args.get('model')}%"))
    products = query.order_by(Product.category, Product.name).all()
    return render_template('products.html', products=products)


@app.route('/product/new', methods=['POST'])
@login_required
@permission_required('增加')
def new_product():
    product = Product(
        name=request.form['name'],
        category=request.form['category'],
        model=request.form.get('model'),
        unit=request.form.get('unit'),
        # 新增字段
        tax_rate=float(request.form['tax_rate']) if request.form.get('tax_rate') else None,
        ref_quantity=float(request.form['ref_quantity']) if request.form.get('ref_quantity') else None,
        ref_unit_price=float(request.form['ref_unit_price']) if request.form.get('ref_unit_price') else None,
        customer_id=get_current_customer_id()  # 新增：自动关联租户
    )
    db.session.add(product)
    db.session.commit()
    flash('产品添加成功', 'success')
    return redirect(url_for('products'))


@app.route('/product/<int:id>/edit', methods=['POST'])
@login_required
@permission_required('修改')
def edit_product(id):
    product = Product.query.get_or_404(id)
    product.name = request.form['name']
    product.category = request.form['category']
    product.model = request.form.get('model')
    product.unit = request.form.get('unit')
    # 新增字段
    product.tax_rate = float(request.form['tax_rate']) if request.form.get('tax_rate') else None
    product.ref_quantity = float(request.form['ref_quantity']) if request.form.get('ref_quantity') else None
    product.ref_unit_price = float(request.form['ref_unit_price']) if request.form.get('ref_unit_price') else None
    db.session.commit()
    flash('产品更新成功', 'success')
    return redirect(url_for('products'))


@app.route('/product/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('删除')
def delete_product(id):
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('产品删除成功', 'success')
    return redirect(url_for('products'))


@app.route('/api/products/search')
@login_required
def search_products():
    query = request.args.get('q', '')
    q = Product.query.filter(Product.name.like(f'%{query}%'))

    # 新增：数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        q = q.filter(Product.customer_id == customer_id)

    products = q.limit(10).all()
    return jsonify([{
        'id': p.id, 'name': p.name, 'category': p.category, 'model': p.model, 'unit': p.unit,
        'tax_rate': p.tax_rate, 'ref_quantity': p.ref_quantity, 'ref_unit_price': p.ref_unit_price
    } for p in products])


# ── 新增：自动更新合同状态（已收付款>=合同总价 且 已开票>=合同总价 则标记已完结）──
def auto_update_contract_status(contract):
    """检查合同是否满足完结条件，并自动更新状态"""
    total_paid = contract.get_total_paid()
    total_invoiced = contract.get_total_invoiced()
    if contract.total_price > 0 and total_paid >= contract.total_price and total_invoiced >= contract.total_price:
        contract.status = '已完结'
    else:
        if contract.status == '已完结':
            contract.status = '进行中'


# ── 新增：将合同中的产品名称同步到 Product 表（新产品则创建，已有则跳过）──
def sync_products_to_table(product_names, models, units, tax_rates, customer_id):
    """将新建/编辑合同中的产品名称同步到产品管理表"""
    for i, pname in enumerate(product_names):
        pname = pname.strip() if pname else ''
        if not pname:
            continue
        existing = Product.query.filter_by(name=pname, customer_id=customer_id).first()
        if not existing:
            model_val = models[i].strip() if i < len(models) and models[i] else None
            unit_val = units[i].strip() if i < len(units) and units[i] else None
            tax_val = None
            if i < len(tax_rates) and tax_rates[i]:
                try:
                    tax_val = float(tax_rates[i])
                except (ValueError, TypeError):
                    tax_val = None
            new_product = Product(
                name=pname,
                category='其他',  # 默认分类，用户可在产品管理页修改
                model=model_val or None,
                unit=unit_val or None,
                tax_rate=tax_val,
                customer_id=customer_id
            )
            db.session.add(new_product)


@app.route('/statistics')
@login_required
def statistics():
    from sqlalchemy import func
    q = Contract.query

    # 新增：租户数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        q = q.filter(Contract.customer_id == customer_id)
    # 筛选
    f_staff = request.args.get('f_staff', '')
    f_customer = request.args.get('f_customer', '')
    f_type = request.args.get('f_type', '')
    f_business = request.args.get('f_business', '')
    f_status = request.args.get('f_status', '')
    f_year = request.args.get('f_year', '')   # 新增：年份筛选
    if f_staff:
        q = q.filter(Contract.project_staff.like(f'%{f_staff}%'))
    if f_customer:
        q = q.filter(Contract.customer_name.like(f'%{f_customer}%'))
    if f_type:
        q = q.filter(Contract.contract_type == f_type)
    if f_business:
        q = q.filter(Contract.business_type == f_business)
    if f_status:
        q = q.filter(Contract.status == f_status)
    if f_year:
        q = q.filter(func.strftime('%Y', Contract.signing_date) == f_year)

    # 新增：f_sheets 参数——控制页面上显示哪些维度的统计表格
    # 默认全部显示；用户勾选后只显示选中的
    # 注意：表单 checkbox 以多值形式发送（f_sheets=staff&f_sheets=customer），用 getlist 读取
    all_sheets = ['staff', 'customer', 'type', 'business', 'status', 'detail']
    selected_sheets = [s for s in request.args.getlist('f_sheets') if s in all_sheets]
    if not selected_sheets:
        selected_sheets = all_sheets  # 默认全部

    stats = {}
    if 'staff' in selected_sheets:
        stats['by_staff'] = q.with_entities(Contract.project_staff, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.project_staff).all()
    if 'customer' in selected_sheets:
        stats['by_customer'] = q.with_entities(Contract.customer_name, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.customer_name).all()
    if 'type' in selected_sheets:
        stats['by_type'] = q.with_entities(Contract.contract_type, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.contract_type).all()
    if 'status' in selected_sheets:
        stats['by_status'] = q.with_entities(Contract.status, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.status).all()
    if 'business' in selected_sheets:
        stats['by_business'] = q.with_entities(Contract.business_type, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.business_type).all()
    # 新增：明细数据
    detail_contracts = []
    if 'detail' in selected_sheets:
        detail_contracts = q.order_by(Contract.signing_date.desc()).all()

    filters = {'f_staff': f_staff, 'f_customer': f_customer, 'f_type': f_type,
               'f_business': f_business, 'f_status': f_status, 'f_year': f_year}

    # 新增：获取可用年份列表
    years_raw = db.session.query(func.strftime('%Y', Contract.signing_date)).filter(
        Contract.signing_date.isnot(None)
    ).distinct().order_by(func.strftime('%Y', Contract.signing_date).desc()).all()
    available_years = [int(y[0]) for y in years_raw if y[0]]

    user_permissions = db.session.get(User, session['user_id']).permissions or ''
    return render_template('statistics.html', stats=stats, filters=filters,
                           available_years=available_years, selected_sheets=selected_sheets,
                           detail_contracts=detail_contracts,
                           user_permissions=user_permissions)


# ── 新增：收付款管理（独立页面）──
@app.route('/payments')
@login_required
def payments_list():
    user_id = session.get('user_id')

    # 检查是否有收付款相关权限
    perms = get_user_function_permissions(user_id)
    has_payment_perm = any('收付款' in p for p in perms)

    if not has_payment_perm and perms != 'all':
        flash('您没有收付款管理权限，请联系管理员', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()

    # 获取所有合同
    q = Contract.query

    if customer_id is not None:
        q = q.filter(Contract.customer_id == customer_id)

    # 筛选参数
    f_contract = request.args.get('f_contract', '')
    f_customer = request.args.get('f_customer', '')
    f_status = request.args.get('f_status', '')

    if f_contract:
        q = q.filter(Contract.project_name.like(f'%{f_contract}%'))
    if f_customer:
        q = q.filter(Contract.customer_name.like(f'%{f_customer}%'))
    if f_status:
        q = q.filter(Contract.status == f_status)

    contracts = q.order_by(Contract.signing_date.desc()).all()

    # 为每个合同获取收付款记录
    contracts_with_payments = []
    for contract in contracts:
        payments = Payment.query.filter_by(contract_id=contract.id).order_by(Payment.payment_date.desc()).all()
        contracts_with_payments.append({
            'contract': contract,
            'payments': payments,
            'total_paid': sum(p.amount for p in payments),
            'unpaid': contract.total_price - sum(p.amount for p in payments)
        })

    filters = {
        'f_contract': f_contract,
        'f_customer': f_customer,
        'f_status': f_status
    }

    return render_template('payments_list.html', contracts_with_payments=contracts_with_payments, filters=filters)


# ── 新增：发票管理（独立页面）──
@app.route('/invoices')
@login_required
def invoices_list():
    user_id = session.get('user_id')

    # 检查是否有发票相关权限
    perms = get_user_function_permissions(user_id)
    has_invoice_perm = any('发票' in p for p in perms)

    if not has_invoice_perm and perms != 'all':
        flash('您没有发票管理权限，请联系管理员', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()

    # 获取所有合同
    q = Contract.query

    if customer_id is not None:
        q = q.filter(Contract.customer_id == customer_id)

    # 筛选参数
    f_contract = request.args.get('f_contract', '')
    f_customer = request.args.get('f_customer', '')
    f_status = request.args.get('f_status', '')

    if f_contract:
        q = q.filter(Contract.project_name.like(f'%{f_contract}%'))
    if f_customer:
        q = q.filter(Contract.customer_name.like(f'%{f_customer}%'))
    if f_status:
        q = q.filter(Contract.status == f_status)

    contracts = q.order_by(Contract.signing_date.desc()).all()

    # 为每个合同获取发票记录
    contracts_with_invoices = []
    for contract in contracts:
        invoices = Invoice.query.filter_by(contract_id=contract.id).order_by(Invoice.received_date.desc()).all()
        total_invoiced = sum(i.amount for i in invoices if i.invoice_status == '已开具')
        contracts_with_invoices.append({
            'contract': contract,
            'invoices': invoices,
            'total_invoiced': total_invoiced,
            'uninvoiced': contract.total_price - total_invoiced
        })

    filters = {
        'f_contract': f_contract,
        'f_customer': f_customer,
        'f_status': f_status
    }

    return render_template('invoices_list.html', contracts_with_invoices=contracts_with_invoices, filters=filters)


# ── 新增：交付管理（独立页面）──
@app.route('/deliveries')
@login_required
def deliveries_list():
    user_id = session.get('user_id')

    # 检查是否有交付相关权限
    perms = get_user_function_permissions(user_id)
    has_delivery_perm = any('交付' in p for p in perms)

    if not has_delivery_perm and perms != 'all':
        flash('您没有交付管理权限，请联系管理员', 'warning')
        return redirect(url_for('index'))

    customer_id = get_current_customer_id()

    # 获取所有合同
    q = Contract.query

    if customer_id is not None:
        q = q.filter(Contract.customer_id == customer_id)

    # 筛选参数
    f_contract = request.args.get('f_contract', '')
    f_customer = request.args.get('f_customer', '')
    f_status = request.args.get('f_status', '')

    if f_contract:
        q = q.filter(Contract.project_name.like(f'%{f_contract}%'))
    if f_customer:
        q = q.filter(Contract.customer_name.like(f'%{f_customer}%'))
    if f_status:
        q = q.filter(Contract.status == f_status)

    contracts = q.order_by(Contract.signing_date.desc()).all()

    # 为每个合同获取交付记录
    contracts_with_deliveries = []
    for contract in contracts:
        deliveries = Delivery.query.filter_by(contract_id=contract.id).order_by(Delivery.delivery_date.desc()).all()
        contracts_with_deliveries.append({
            'contract': contract,
            'deliveries': deliveries,
            'delivery_count': len(deliveries)
        })

    filters = {
        'f_contract': f_contract,
        'f_customer': f_customer,
        'f_status': f_status
    }

    return render_template('deliveries_list.html', contracts_with_deliveries=contracts_with_deliveries, filters=filters)


# ── 统计分析导出 Excel ──
@app.route('/statistics/export')
@login_required
@permission_required('导出EXCEL')
def export_statistics():
    from sqlalchemy import func
    q = Contract.query

    # 新增：租户数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        q = q.filter(Contract.customer_id == customer_id)
    f_staff = request.args.get('f_staff', '')
    f_customer = request.args.get('f_customer', '')
    f_type = request.args.get('f_type', '')
    f_business = request.args.get('f_business', '')
    f_status = request.args.get('f_status', '')
    f_year = request.args.get('f_year', '')
    if f_staff:
        q = q.filter(Contract.project_staff.like(f'%{f_staff}%'))
    if f_customer:
        q = q.filter(Contract.customer_name.like(f'%{f_customer}%'))
    if f_type:
        q = q.filter(Contract.contract_type == f_type)
    if f_business:
        q = q.filter(Contract.business_type == f_business)
    if f_status:
        q = q.filter(Contract.status == f_status)
    if f_year:
        q = q.filter(func.strftime('%Y', Contract.signing_date) == f_year)

    # 按勾选的sheets参数决定导出哪些sheet
    sheets = request.args.get('sheets', 'staff,customer,type,business,status').split(',')
    layout = request.args.get('layout', 'vertical')  # vertical 或 horizontal

    # 新增：明细数据导出
    detail_rows = []
    detail_spans = []  # (start, count) 每个合同的行范围
    if 'detail' in sheets:
        detail_contracts = q.order_by(Contract.signing_date.desc()).all()
        DETAIL_MERGE_COLS = [0,1,2,4,6,7,8,9,10,11,12,13,14,15]  # 合同级列索引
        for c in detail_contracts:
            products = ContractProduct.query.filter_by(contract_id=c.id).all()
            start = len(detail_rows)
            base = [c.contract_number or '', c.customer_name, c.project_name,
                    '', c.total_price, '',
                    c.contract_type or '', c.business_type or '',
                    c.project_staff or '', c.sales_staff or '',
                    str(c.signing_date) if c.signing_date else '', c.status or '',
                    c.get_total_paid(), c.get_unpaid_amount(),
                    c.get_total_invoiced(), c.get_uninvoiced_amount()]
            if products:
                for cp in products:
                    row = list(base)
                    row[3] = cp.product_name or ''
                    row[5] = cp.tax_rate
                    detail_rows.append(row)
            else:
                row = list(base)
                row[3] = c.product_name or ''
                row[5] = c.tax_rate
                detail_rows.append(row)
            detail_spans.append((start, len(detail_rows) - start))

    # 收集各维度数据
    blocks = []  # [(col_name, data), ...]
    if 'staff' in sheets:
        blocks.append(('项目负责人', q.with_entities(Contract.project_staff, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.project_staff).all()))
    if 'customer' in sheets:
        blocks.append(('客户名称', q.with_entities(Contract.customer_name, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.customer_name).all()))
    if 'type' in sheets:
        blocks.append(('合同类型', q.with_entities(Contract.contract_type, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.contract_type).all()))
    if 'business' in sheets:
        blocks.append(('业务类型', q.with_entities(Contract.business_type, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.business_type).all()))
    if 'status' in sheets:
        blocks.append(('履约状态', q.with_entities(Contract.status, func.count(Contract.id), func.sum(Contract.total_price)).group_by(Contract.status).all()))

    import openpyxl
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '统计分析'

    if layout == 'multisheet':
        # 方案C：每个维度单独一个 sheet
        first = True
        for col_name, data in blocks:
            ws = wb.active if first else wb.create_sheet()
            ws.title = col_name
            first = False
            ws.cell(row=1, column=1, value=col_name)
            ws.cell(row=1, column=2, value='合同数量')
            ws.cell(row=1, column=3, value='合同总额')
            for i, r in enumerate(data, start=2):
                ws.cell(row=i, column=1, value=r[0])
                ws.cell(row=i, column=2, value=r[1])
                ws.cell(row=i, column=3, value=float(r[2]) if r[2] else 0)
    elif layout == 'horizontal':
        # 方案B：横向排列，每维度占3列，列间空1列
        col = 1
        for col_name, data in blocks:
            ws.cell(row=1, column=col, value=col_name)
            ws.cell(row=1, column=col+1, value='合同数量')
            ws.cell(row=1, column=col+2, value='合同总额')
            for i, r in enumerate(data, start=2):
                ws.cell(row=i, column=col, value=r[0])
                ws.cell(row=i, column=col+1, value=r[1])
                ws.cell(row=i, column=col+2, value=float(r[2]) if r[2] else 0)
            col += 4  # 3列数据 + 1列空白
    else:
        # 方案A：纵向排列
        row = 1
        for col_name, data in blocks:
            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value='合同数量')
            ws.cell(row=row, column=3, value='合同总额')
            row += 1
            for r in data:
                ws.cell(row=row, column=1, value=r[0])
                ws.cell(row=row, column=2, value=r[1])
                ws.cell(row=row, column=3, value=float(r[2]) if r[2] else 0)
                row += 1
            row += 1  # 空行分隔

    wb.save(buf)
    buf.seek(0)
    # 新增：如有明细数据，写入明细sheet
    if detail_rows:
        import openpyxl as _opx
        buf2 = io.BytesIO()
        buf.seek(0)
        wb2 = _opx.load_workbook(buf)
        ws_detail = wb2.create_sheet(title='明细数据')
        detail_headers = ['合同编号','客户名称','项目名称','产品名称','合同总价','发票税率',
                          '合同类型','业务类型','项目负责人','销售人员','签订日期','状态',
                          '已收付款','未收付款','已开票','未开票']
        for ci, h in enumerate(detail_headers, 1):
            ws_detail.cell(row=1, column=ci, value=h)
        for ri, row in enumerate(detail_rows, 2):
            for ci, val in enumerate(row, 1):
                ws_detail.cell(row=ri, column=ci, value=val)
        # 合并同一合同的合同级列
        from openpyxl.utils import get_column_letter as _gcl
        for start, span in detail_spans:
            if span > 1:
                for col_idx in DETAIL_MERGE_COLS:
                    cl = _gcl(col_idx + 1)
                    ws_detail.merge_cells(f'{cl}{start+2}:{cl}{start+span+1}')
        wb2.save(buf2)
        buf2.seek(0)
        filename = f"统计导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return send_file(buf2, download_name=filename, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    buf.seek(0)
    filename = f"统计导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# 新增：带业务类型的新建合同路由
@app.route('/contract/new/<business_type>', methods=['GET', 'POST'])
@login_required
@permission_required('增加')
def new_contract_with_type(business_type):
    """新建合同（指定业务类型：采购/销售）"""
    if business_type not in ['采购', '销售']:
        flash('业务类型错误', 'warning')
        return redirect(url_for('index'))

    if request.method == 'POST':
        # 获取当前用户的租户ID
        customer_id = get_current_customer_id()

        # 创建合同主记录
        contract = Contract(
            contract_number=request.form.get('contract_number'),
            customer_name=request.form['customer_name'],
            project_name=request.form['project_name'],
            total_price=float(request.form['total_price']),
            contract_type=request.form.get('contract_type'),
            project_staff=request.form.get('project_staff'),
            sales_staff=request.form.get('sales_staff'),
            business_type=business_type,  # 使用URL参数指定的业务类型
            signing_date=datetime.strptime(request.form['signing_date'], '%Y-%m-%d').date() if request.form.get('signing_date') else None,
            customer_id=customer_id,
            created_by=session.get('username')
        )

        # 新增：保存所属组织
        org_id = request.form.get('organization_id')
        if org_id and org_id.strip():
            contract.organization_id = int(org_id)

        # 处理合同文件上传
        if 'contract_file' in request.files:
            file = request.files['contract_file']
            if file.filename:
                filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                contract.file_path = filename

        # 自动同步客户信息
        if not Customer.query.filter_by(name=contract.customer_name, customer_id=customer_id).first():
            db.session.add(Customer(name=contract.customer_name, customer_id=customer_id))

        db.session.add(contract)
        db.session.flush()

        # 处理多产品数据
        product_names = request.form.getlist('products[product_name][]')
        contract_types = request.form.getlist('products[contract_type][]')
        product_types = request.form.getlist('products[product_type][]')
        models = request.form.getlist('products[model][]')
        units = request.form.getlist('products[unit][]')
        quantities = request.form.getlist('products[quantity][]')
        unit_prices = request.form.getlist('products[unit_price][]')
        subtotals = request.form.getlist('products[subtotal][]')
        tax_rates = request.form.getlist('products[tax_rate][]')

        for i in range(len(product_names)):
            if product_names[i].strip():
                cp = ContractProduct(
                    contract_id=contract.id,
                    product_name=product_names[i].strip() or None,
                    contract_type=contract_types[i] if i < len(contract_types) else None,
                    product_type=product_types[i].strip() if i < len(product_types) and product_types[i].strip() else None,
                    model=models[i].strip() if i < len(models) and models[i].strip() else None,
                    unit=units[i].strip() if i < len(units) and units[i].strip() else None,
                    quantity=float(quantities[i]) if i < len(quantities) and quantities[i] else None,
                    unit_price=float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else None,
                    subtotal=float(subtotals[i]) if i < len(subtotals) and subtotals[i] else None,
                    tax_rate=float(tax_rates[i]) if i < len(tax_rates) and tax_rates[i] else None
                )
                db.session.add(cp)

        # 将产品名称同步到产品管理表
        sync_products_to_table(product_names, models, units, tax_rates, customer_id)

        db.session.commit()
        flash(f'{business_type}合同创建成功', 'success')
        return redirect(url_for('contract_archive', business_type=business_type))

    # GET请求：数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        customers_list = Customer.query.filter_by(customer_id=customer_id).order_by(Customer.name).all()
        products_list = Product.query.filter_by(customer_id=customer_id).order_by(Product.name).all()
        organizations_list = Organization.query.filter_by(customer_id=customer_id).order_by(Organization.name).all()
    else:
        customers_list = Customer.query.order_by(Customer.name).all()
        products_list = Product.query.order_by(Product.name).all()
        organizations_list = []

    return render_template('contract_form.html',
                         customers_list=customers_list,
                         products_list=products_list,
                         organizations_list=organizations_list,
                         business_type=business_type)


# 新增：合同档案路由（按业务类型筛选）
@app.route('/contract/archive/<business_type>')
@login_required
def contract_archive(business_type):
    """合同档案（采购/销售）"""
    if business_type not in ['采购', '销售']:
        flash('业务类型错误', 'warning')
        return redirect(url_for('index'))

    from sqlalchemy import func
    query = Contract.query.filter(Contract.business_type == business_type)

    # 数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        query = query.filter(Contract.customer_id == customer_id)

    # 新增：数据权限过滤
    user_id = session.get('user_id')
    if user_id and customer_id is not None:
        # 新增：读取"包含已完结合同"选项
        include_finished = request.args.get('include_finished') == '1'
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id, include_finished)
        # None 表示可访问所有合同，不需要过滤
        if accessible_ids is not None:
            if not accessible_ids:
                # 空列表表示没有可访问的合同
                contracts = []
                stats = {
                    'count': 0,
                    'total_price': 0,
                    'total_paid': 0,
                    'total_unpaid': 0,
                    'total_invoiced': 0,
                    'total_uninvoiced': 0,
                }
                user = db.session.get(User, user_id)
                return render_template('contract_archive.html',
                                     contracts=[],
                                     stats=stats,
                                     available_years=[],
                                     alerts=[],
                                     user_permissions=get_user_permissions_string(user_id),
                                     is_tenant_user=user.customer_id is not None if user else False,
                                     business_type=business_type,
                                     page=1,
                                     per_page=10,
                                     total_pages=0,
                                     total_contracts=0)
            else:
                # 过滤可访问的合同
                query = query.filter(Contract.id.in_(accessible_ids))

    # 原有筛选条件
    if request.args.get('project_staff'):
        query = query.filter(Contract.project_staff.like(f"%{request.args.get('project_staff')}%"))
    if request.args.get('customer_name'):
        query = query.filter(Contract.customer_name.like(f"%{request.args.get('customer_name')}%"))

    # 合同类型筛选
    if request.args.get('contract_type'):
        contract_type = request.args.get('contract_type')
        query = query.join(ContractProduct, Contract.id == ContractProduct.contract_id).filter(
            ContractProduct.contract_type == contract_type
        ).distinct()

    if request.args.get('status'):
        query = query.filter(Contract.status == request.args.get('status'))

    # 签订年份筛选
    if request.args.get('signing_year'):
        query = query.filter(func.strftime('%Y', Contract.signing_date) == request.args.get('signing_year'))

    contracts = query.order_by(Contract.created_at.desc()).all()

    # 发票状态筛选
    if request.args.get('invoice_status'):
        filtered = []
        for contract in contracts:
            has_issued = any(i.invoice_status == '已开具' for i in contract.invoices)
            if request.args.get('invoice_status') == '已开具' and has_issued:
                filtered.append(contract)
            elif request.args.get('invoice_status') == '未开具' and not has_issued:
                filtered.append(contract)
        contracts = filtered

    # 预警筛选
    if request.args.get('alert') == 'yes':
        contracts = [c for c in contracts if c.get_unpaid_amount() > 0 or c.get_uninvoiced_amount() > 0]

    # 统计数据
    stats = {
        'count': len(contracts),
        'total_price': sum(c.total_price for c in contracts),
        'total_paid': sum(c.get_total_paid() for c in contracts),
        'total_unpaid': sum(c.get_unpaid_amount() for c in contracts),
        'total_invoiced': sum(c.get_total_invoiced() for c in contracts),
        'total_uninvoiced': sum(c.get_uninvoiced_amount() for c in contracts),
    }

    # 获取可用年份
    all_contracts = Contract.query.filter(Contract.business_type == business_type)
    if customer_id is not None:
        all_contracts = all_contracts.filter(Contract.customer_id == customer_id)
    available_years = sorted(set(
        c.signing_date.year for c in all_contracts.all() if c.signing_date
    ), reverse=True)

    # 收付款预警
    alerts = []
    user = db.session.get(User, session['user_id'])
    user_permissions = get_user_permissions_string(session['user_id'])
    is_tenant_user = user.customer_id is not None if user else False

    # 新增：分页功能
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)

    total_contracts = len(contracts)
    total_pages = (total_contracts + per_page - 1) // per_page

    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    contracts_page = contracts[start_idx:end_idx]

    return render_template('contract_archive.html',
                         contracts=contracts_page,
                         stats=stats,
                         available_years=available_years,
                         alerts=alerts,
                         user_permissions=user_permissions,
                         is_tenant_user=is_tenant_user,
                         business_type=business_type,
                         page=page,
                         per_page=per_page,
                         total_pages=total_pages,
                         total_contracts=total_contracts)


@app.route('/contract/new', methods=['GET', 'POST'])
@login_required
@permission_required('增加')
def new_contract():
    if request.method == 'POST':
        # 新增：获取当前用户的租户ID
        customer_id = get_current_customer_id()

        # 创建合同主记录
        contract = Contract(
            contract_number=request.form.get('contract_number'),
            customer_name=request.form['customer_name'],
            project_name=request.form['project_name'],
            total_price=float(request.form['total_price']),
            contract_type=request.form.get('contract_type'),
            project_staff=request.form.get('project_staff'),
            sales_staff=request.form.get('sales_staff'),
            business_type=request.form.get('business_type', '销售'),
            signing_date=datetime.strptime(request.form['signing_date'], '%Y-%m-%d').date() if request.form.get('signing_date') else None,
            customer_id=customer_id,  # 新增：关联租户
            created_by=session.get('username')  # 新增：记录创建人
        )

        # 新增：保存所属组织
        org_id = request.form.get('organization_id')
        if org_id and org_id.strip():
            contract.organization_id = int(org_id)

        # 处理合同文件上传
        if 'contract_file' in request.files:
            file = request.files['contract_file']
            if file.filename:
                filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                contract.file_path = filename

        # 自动同步客户信息
        if not Customer.query.filter_by(name=contract.customer_name, customer_id=customer_id).first():
            db.session.add(Customer(name=contract.customer_name, customer_id=customer_id))

        db.session.add(contract)
        db.session.flush()  # 获取contract.id

        # 新增：处理多产品数据
        product_names = request.form.getlist('products[product_name][]')
        contract_types = request.form.getlist('products[contract_type][]')
        product_types = request.form.getlist('products[product_type][]')
        models = request.form.getlist('products[model][]')
        units = request.form.getlist('products[unit][]')
        quantities = request.form.getlist('products[quantity][]')
        unit_prices = request.form.getlist('products[unit_price][]')
        subtotals = request.form.getlist('products[subtotal][]')
        tax_rates = request.form.getlist('products[tax_rate][]')

        # 保存多个产品到 ContractProduct 表
        for i in range(len(product_names)):
            if product_names[i].strip():  # 只保存非空产品
                cp = ContractProduct(
                    contract_id=contract.id,
                    product_name=product_names[i].strip() or None,
                    contract_type=contract_types[i] if i < len(contract_types) else None,
                    product_type=product_types[i].strip() if i < len(product_types) and product_types[i].strip() else None,
                    model=models[i].strip() if i < len(models) and models[i].strip() else None,
                    unit=units[i].strip() if i < len(units) and units[i].strip() else None,
                    quantity=float(quantities[i]) if i < len(quantities) and quantities[i] else None,
                    unit_price=float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else None,
                    subtotal=float(subtotals[i]) if i < len(subtotals) and subtotals[i] else None,
                    tax_rate=float(tax_rates[i]) if i < len(tax_rates) and tax_rates[i] else None
                )
                db.session.add(cp)

        # 新增：将产品名称同步到产品管理表（仅新产品）
        sync_products_to_table(product_names, models, units, tax_rates, customer_id)

        db.session.commit()
        flash('合同创建成功', 'success')
        return redirect(url_for('index'))

    # GET请求：数据隔离
    customer_id = get_current_customer_id()
    if customer_id is not None:
        customers_list = Customer.query.filter_by(customer_id=customer_id).order_by(Customer.name).all()
        products_list = Product.query.filter_by(customer_id=customer_id).order_by(Product.name).all()
        organizations_list = Organization.query.filter_by(customer_id=customer_id).order_by(Organization.name).all()
    else:
        customers_list = Customer.query.order_by(Customer.name).all()
        products_list = Product.query.order_by(Product.name).all()
        organizations_list = []

    return render_template('contract_form.html', customers_list=customers_list, products_list=products_list, organizations_list=organizations_list)


@app.route('/contract/<int:id>')
@login_required
def view_contract(id):
    contract = Contract.query.get_or_404(id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    user_id = session.get('user_id')
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and id not in accessible_ids:
            flash('您没有权限查看该合同', 'warning')
            return redirect(url_for('index'))

    return render_template('contract_detail.html', contract=contract, view_only=True)


@app.route('/contract/<int:id>/manage')
@login_required
def manage_contract(id):
    contract = Contract.query.get_or_404(id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        app.logger.warning(f"租户隔离检查失败: user_customer={customer_id}, contract_customer={contract.customer_id}")
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '修改', contract):
        flash('您没有权限管理该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        app.logger.info(f"数据权限检查: user_id={user_id}, contract_id={id}, accessible_ids={'全部' if accessible_ids is None else accessible_ids}")
        if accessible_ids is not None and id not in accessible_ids:
            app.logger.warning(f"数据权限检查失败: contract_id={id} 不在可访问列表中")
            flash('您没有权限管理该合同', 'warning')
            return redirect(url_for('index'))

    return render_template('contract_detail.html', contract=contract, view_only=False)


@app.route('/contract/<int:id>/delete', methods=['POST'])
@login_required
def delete_contract(id):
    contract = Contract.query.get_or_404(id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '删除', contract):
        flash('您没有权限删除该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and id not in accessible_ids:
            flash('您没有权限删除该合同', 'warning')
            return redirect(url_for('index'))

    db.session.delete(contract)
    db.session.commit()
    flash('合同已删除', 'success')
    return redirect(url_for('index'))


# 新增：批量删除合同
@app.route('/contract/batch_delete', methods=['POST'])
@login_required
@permission_required('删除')
def batch_delete_contracts():
    """批量删除合同"""
    contract_ids = request.form.getlist('contract_ids')

    if not contract_ids:
        flash('未选择任何合同', 'warning')
        return redirect(url_for('index'))

    # 数据隔离检查
    customer_id = get_current_customer_id()
    user_id = session.get('user_id')

    # 获取用户可访问的合同ID
    accessible_ids = None
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)

    deleted_count = 0
    for contract_id in contract_ids:
        contract = db.session.get(Contract, contract_id)
        if contract:
            # 检查租户权限
            if customer_id is not None and contract.customer_id != customer_id:
                continue  # 跳过其他租户的合同

            # 检查数据权限
            if accessible_ids is not None and int(contract_id) not in accessible_ids:
                continue  # 跳过无权限的合同

            db.session.delete(contract)
            deleted_count += 1

    db.session.commit()

    if deleted_count > 0:
        flash(f'成功删除 {deleted_count} 条合同', 'success')
    else:
        flash('没有合同被删除', 'warning')

    return redirect(url_for('index'))


@app.route('/contract/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_contract(id):
    contract = Contract.query.get_or_404(id)

    # 新增：数据隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足：无法访问其他租户的合同', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '修改', contract):
        flash('您没有权限编辑该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and id not in accessible_ids:
            flash('您没有权限编辑该合同', 'warning')
            return redirect(url_for('index'))

    if request.method == 'POST':
        contract.customer_name = request.form['customer_name']
        contract.project_name = request.form['project_name']
        contract.contract_number = request.form.get('contract_number')
        contract.contract_type = request.form.get('contract_type')
        contract.total_price = float(request.form['total_price'])
        contract.project_staff = request.form.get('project_staff')
        contract.sales_staff = request.form.get('sales_staff')
        contract.status = request.form.get('status')
        contract.business_type = request.form.get('business_type', '销售')
        contract.signing_date = datetime.strptime(request.form['signing_date'], '%Y-%m-%d').date() if request.form.get('signing_date') else None

        # 新增：保存所属组织
        org_id = request.form.get('organization_id')
        if org_id and org_id.strip():
            contract.organization_id = int(org_id)
        else:
            contract.organization_id = None

        if 'contract_file' in request.files:
            file = request.files['contract_file']
            if file.filename:
                filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                contract.file_path = filename

        # 新增：更新多产品数据 - 先删除旧的，再添加新的
        ContractProduct.query.filter_by(contract_id=contract.id).delete()

        product_names = request.form.getlist('products[product_name][]')
        contract_types = request.form.getlist('products[contract_type][]')
        product_types = request.form.getlist('products[product_type][]')
        models = request.form.getlist('products[model][]')
        units = request.form.getlist('products[unit][]')
        quantities = request.form.getlist('products[quantity][]')
        unit_prices = request.form.getlist('products[unit_price][]')
        subtotals = request.form.getlist('products[subtotal][]')
        tax_rates = request.form.getlist('products[tax_rate][]')

        for i in range(len(product_names)):
            if product_names[i].strip():
                cp = ContractProduct(
                    contract_id=contract.id,
                    product_name=product_names[i].strip() or None,
                    contract_type=contract_types[i] if i < len(contract_types) else None,
                    product_type=product_types[i].strip() if i < len(product_types) and product_types[i].strip() else None,
                    model=models[i].strip() if i < len(models) and models[i].strip() else None,
                    unit=units[i].strip() if i < len(units) and units[i].strip() else None,
                    quantity=float(quantities[i]) if i < len(quantities) and quantities[i] else None,
                    unit_price=float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else None,
                    subtotal=float(subtotals[i]) if i < len(subtotals) and subtotals[i] else None,
                    tax_rate=float(tax_rates[i]) if i < len(tax_rates) and tax_rates[i] else None
                )
                db.session.add(cp)

        # 新增：将产品名称同步到产品管理表（仅新产品）
        sync_products_to_table(product_names, models, units, tax_rates, customer_id)

        db.session.flush()
        auto_update_contract_status(contract)
        db.session.commit()
        flash('合同更新成功', 'success')
        return redirect(url_for('view_contract', id=id))

    # GET请求：数据隔离
    if customer_id is not None:
        customers_list = Customer.query.filter_by(customer_id=customer_id).order_by(Customer.name).all()
        products_list = Product.query.filter_by(customer_id=customer_id).order_by(Product.name).all()
        organizations_list = Organization.query.filter_by(customer_id=customer_id).order_by(Organization.name).all()
    else:
        customers_list = Customer.query.order_by(Customer.name).all()
        products_list = Product.query.order_by(Product.name).all()
        organizations_list = []

    return render_template('contract_form.html', contract=contract, customers_list=customers_list, products_list=products_list, organizations_list=organizations_list)


@app.route('/contract/<int:id>/payment', methods=['GET', 'POST'])
@login_required
def add_payment(id):
    contract = Contract.query.get_or_404(id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    # 检查是否是项目负责人
    is_project_manager = contract and contract.project_staff and session.get('username') in contract.project_staff

    # 检查 '增加-收付款' 权限，如果没有则检查旧格式 '增加' 权限
    has_add_payment_perm = has_permission(user_id, '增加-收付款') or has_permission(user_id, '增加')

    if not is_project_manager and not has_add_payment_perm:
        flash('您没有权限添加收付款记录', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    if request.method == 'POST':
        payment = Payment(
            contract_id=id,
            amount=float(request.form['amount']),
            payment_date=datetime.strptime(request.form['payment_date'], '%Y-%m-%d').date(),
            payment_type=request.form.get('payment_type'),
            note=request.form.get('note')
        )

        if 'receipt_file' in request.files:
            files = request.files.getlist('receipt_file')
            saved = []
            for file in files:
                if file.filename:
                    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{file.filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    saved.append(filename)
            if saved:
                payment.receipt_file = ','.join(saved)

        db.session.add(payment)
        db.session.commit()
        # 新增：自动更新合同状态
        contract = db.session.get(Contract, id)
        if contract:
            auto_update_contract_status(contract)
            db.session.commit()
        flash('收付款记录添加成功', 'success')

        # 判断来源页面
        referer = request.referrer or ''
        if 'payments' in referer:
            return redirect(url_for('payments_list'))
        else:
            return redirect(url_for('view_contract', id=id))

    # GET 请求：显示添加表单
    return render_template('add_payment.html', contract=contract)


@app.route('/contract/<int:id>/delivery', methods=['GET', 'POST'])
@login_required
def add_delivery(id):
    contract = Contract.query.get_or_404(id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    # 检查是否是项目负责人
    is_project_manager = contract and contract.project_staff and session.get('username') in contract.project_staff

    # 检查 '增加-交付' 权限，如果没有则检查旧格式 '增加' 权限
    has_add_delivery_perm = has_permission(user_id, '增加-交付') or has_permission(user_id, '增加')

    if not is_project_manager and not has_add_delivery_perm:
        flash('您没有权限添加交付记录', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    if request.method == 'POST':
        delivery = Delivery(
            contract_id=id,
            delivery_date=datetime.strptime(request.form['delivery_date'], '%Y-%m-%d').date(),
            content=request.form.get('content'),
            note=request.form.get('note')
        )

        if 'delivery_file' in request.files:
            files = request.files.getlist('delivery_file')
            saved = []
            for file in files:
                if file.filename:
                    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{file.filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    saved.append(filename)
            if saved:
                delivery.delivery_file = ','.join(saved)

        db.session.add(delivery)
        db.session.commit()
        flash('交付记录添加成功', 'success')

        # 判断来源页面
        referer = request.referrer or ''
        if 'deliveries' in referer:
            return redirect(url_for('deliveries_list'))
        else:
            return redirect(url_for('view_contract', id=id))

    # GET 请求：显示添加表单
    return render_template('add_delivery.html', contract=contract)


@app.route('/contract/<int:id>/invoice', methods=['GET', 'POST'])
@login_required
def add_invoice(id):
    contract = Contract.query.get_or_404(id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    # 检查是否是项目负责人
    is_project_manager = contract and contract.project_staff and session.get('username') in contract.project_staff

    # 检查 '增加-发票' 权限，如果没有则检查旧格式 '增加' 权限
    has_add_invoice_perm = has_permission(user_id, '增加-发票') or has_permission(user_id, '增加')

    if not is_project_manager and not has_add_invoice_perm:
        flash('您没有权限添加发票记录', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    if request.method == 'POST':
        invoice_number = request.form.get('invoice_number', '').strip()
        # 发票号重复检测
        if invoice_number:
            existing = Invoice.query.filter_by(invoice_number=invoice_number).first()
            if existing:
                flash(f'发票号"{invoice_number}"已存在，请核对重新输入', 'warning')
                return redirect(url_for('view_contract', id=id))

        invoice = Invoice(
            contract_id=id,
            amount=float(request.form['amount']),
            received_date=datetime.strptime(request.form['received_date'], '%Y-%m-%d').date(),
            invoice_number=invoice_number or None,
            note=request.form.get('note'),
            invoice_status=request.form.get('invoice_status', '未开具'),
            invoice_type=request.form.get('invoice_type', '普票'),
        )

        if 'invoice_file' in request.files:
            files = request.files.getlist('invoice_file')
            saved = []
            for file in files:
                if file.filename:
                    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{file.filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    saved.append(filename)
            if saved:
                invoice.invoice_file = ','.join(saved)

        db.session.add(invoice)
        db.session.commit()
        # 新增：自动更新合同状态
        contract = db.session.get(Contract, id)
        if contract:
            auto_update_contract_status(contract)
            db.session.commit()
        flash('发票记录添加成功', 'success')

        # 判断来源页面
        referer = request.referrer or ''
        if 'invoices' in referer:
            return redirect(url_for('invoices_list'))
        else:
            return redirect(url_for('view_contract', id=id))

    # GET 请求：显示添加表单
    return render_template('add_invoice.html', contract=contract)


@app.route('/payment/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
def edit_payment(pid):
    payment = Payment.query.get_or_404(pid)
    contract = Contract.query.get_or_404(payment.contract_id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '修改', contract):
        flash('您没有权限操作该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and contract.id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    if request.method == 'POST':
        payment.amount = float(request.form['amount'])
        payment.payment_date = datetime.strptime(request.form['payment_date'], '%Y-%m-%d').date()
        payment.payment_type = request.form.get('payment_type')
        payment.note = request.form.get('note')

        # 处理新上传的文件
        if 'receipt_file' in request.files:
            files = request.files.getlist('receipt_file')
            saved = []
            for file in files:
                if file.filename:
                    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{file.filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    saved.append(filename)
            if saved:
                # 如果已有文件，追加；如果没有，直接设置
                if payment.receipt_file:
                    payment.receipt_file = payment.receipt_file + ',' + ','.join(saved)
                else:
                    payment.receipt_file = ','.join(saved)

        db.session.commit()
        # 自动更新合同状态
        auto_update_contract_status(contract)
        db.session.commit()
        flash('收付款记录更新成功', 'success')
        return redirect(url_for('view_contract', id=contract.id))

    return render_template('edit_payment.html', payment=payment, contract=contract)


@app.route('/payment/<int:pid>/delete', methods=['POST'])
@login_required
def delete_payment(pid):
    payment = Payment.query.get_or_404(pid)
    contract = Contract.query.get_or_404(payment.contract_id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '删除', contract):
        flash('您没有权限操作该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and contract.id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    contract_id = payment.contract_id
    if payment.receipt_file:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], payment.receipt_file))
        except Exception:
            pass
    db.session.delete(payment)
    db.session.flush()
    contract = db.session.get(Contract, contract_id)
    if contract:
        auto_update_contract_status(contract)
    db.session.commit()
    flash('收付款记录已删除', 'success')
    return redirect(url_for('view_contract', id=contract_id))


@app.route('/payment/<int:pid>/delete_file', methods=['POST'])
@login_required
def delete_payment_file(pid):
    payment = Payment.query.get_or_404(pid)
    if payment.receipt_file:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], payment.receipt_file))
        except Exception:
            pass
        payment.receipt_file = None
        db.session.commit()
    return redirect(url_for('view_contract', id=payment.contract_id))


@app.route('/delivery/<int:did>/edit', methods=['GET', 'POST'])
@login_required
def edit_delivery(did):
    delivery = Delivery.query.get_or_404(did)
    contract = Contract.query.get_or_404(delivery.contract_id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '修改', contract):
        flash('您没有权限操作该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and contract.id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    if request.method == 'POST':
        delivery.delivery_date = datetime.strptime(request.form['delivery_date'], '%Y-%m-%d').date()
        delivery.content = request.form.get('content')
        delivery.note = request.form.get('note')

        # 处理新上传的文件
        if 'delivery_file' in request.files:
            files = request.files.getlist('delivery_file')
            saved = []
            for file in files:
                if file.filename:
                    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{file.filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    saved.append(filename)
            if saved:
                # 如果已有文件，追加；如果没有，直接设置
                if delivery.delivery_file:
                    delivery.delivery_file = delivery.delivery_file + ',' + ','.join(saved)
                else:
                    delivery.delivery_file = ','.join(saved)

        db.session.commit()
        flash('交付记录更新成功', 'success')
        return redirect(url_for('view_contract', id=contract.id))

    return render_template('edit_delivery.html', delivery=delivery, contract=contract)


@app.route('/delivery/<int:did>/delete', methods=['POST'])
@login_required
def delete_delivery(did):
    delivery = Delivery.query.get_or_404(did)
    contract = Contract.query.get_or_404(delivery.contract_id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '删除', contract):
        flash('您没有权限操作该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and contract.id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    contract_id = delivery.contract_id
    if delivery.delivery_file:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], delivery.delivery_file))
        except Exception:
            pass
    db.session.delete(delivery)
    db.session.commit()
    flash('交付记录已删除', 'success')
    return redirect(url_for('view_contract', id=contract_id))


@app.route('/delivery/<int:did>/delete_file', methods=['POST'])
@login_required
def delete_delivery_file(did):
    delivery = Delivery.query.get_or_404(did)
    if delivery.delivery_file:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], delivery.delivery_file))
        except Exception:
            pass
        delivery.delivery_file = None
        db.session.commit()
    return redirect(url_for('view_contract', id=delivery.contract_id))


@app.route('/invoice/<int:iid>/edit', methods=['GET', 'POST'])
@login_required
def edit_invoice(iid):
    invoice = Invoice.query.get_or_404(iid)
    contract = Contract.query.get_or_404(invoice.contract_id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '修改', contract):
        flash('您没有权限操作该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and contract.id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    if request.method == 'POST':
        invoice_number = request.form.get('invoice_number', '').strip()
        # 发票号重复检测（排除当前记录）
        if invoice_number:
            existing = Invoice.query.filter_by(invoice_number=invoice_number).filter(Invoice.id != iid).first()
            if existing:
                flash(f'发票号"{invoice_number}"已存在，请核对重新输入', 'warning')
                return redirect(url_for('edit_invoice', iid=iid))

        invoice.amount = float(request.form['amount'])
        invoice.received_date = datetime.strptime(request.form['received_date'], '%Y-%m-%d').date()
        invoice.invoice_number = invoice_number or None
        invoice.note = request.form.get('note')
        invoice.invoice_status = request.form.get('invoice_status', '未开具')
        invoice.invoice_type = request.form.get('invoice_type', '普票')

        # 处理新上传的文件
        if 'invoice_file' in request.files:
            files = request.files.getlist('invoice_file')
            saved = []
            for file in files:
                if file.filename:
                    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{file.filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    saved.append(filename)
            if saved:
                # 如果已有文件，追加；如果没有，直接设置
                if invoice.invoice_file:
                    invoice.invoice_file = invoice.invoice_file + ',' + ','.join(saved)
                else:
                    invoice.invoice_file = ','.join(saved)

        db.session.commit()
        # 自动更新合同状态
        auto_update_contract_status(contract)
        db.session.commit()
        flash('发票记录更新成功', 'success')
        return redirect(url_for('view_contract', id=contract.id))

    return render_template('edit_invoice.html', invoice=invoice, contract=contract)


@app.route('/invoice/<int:iid>/delete', methods=['POST'])
@login_required
def delete_invoice(iid):
    invoice = Invoice.query.get_or_404(iid)
    contract = Contract.query.get_or_404(invoice.contract_id)

    # 租户隔离检查
    customer_id = get_current_customer_id()
    if customer_id is not None and contract.customer_id != customer_id:
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    # 功能权限检查（项目负责人自动拥有权限）
    user_id = session.get('user_id')
    if not has_permission_for_contract(user_id, '删除', contract):
        flash('您没有权限操作该合同', 'warning')
        return redirect(url_for('index'))

    # 数据权限检查
    if user_id and customer_id is not None:
        accessible_ids = get_user_accessible_contract_ids(user_id, customer_id)
        if accessible_ids is not None and contract.id not in accessible_ids:
            flash('您没有权限操作该合同', 'warning')
            return redirect(url_for('index'))

    contract_id = invoice.contract_id
    if invoice.invoice_file:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], invoice.invoice_file))
        except Exception:
            pass
    db.session.delete(invoice)
    db.session.flush()
    contract = db.session.get(Contract, contract_id)
    if contract:
        auto_update_contract_status(contract)
    db.session.commit()
    flash('发票记录已删除', 'success')
    return redirect(url_for('view_contract', id=contract_id))


@app.route('/invoice/<int:iid>/delete_file', methods=['POST'])
@login_required
def delete_invoice_file(iid):
    invoice = Invoice.query.get_or_404(iid)
    if invoice.invoice_file:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], invoice.invoice_file))
        except Exception:
            pass
        invoice.invoice_file = None
        db.session.commit()
    return redirect(url_for('view_contract', id=invoice.contract_id))


@app.route('/import', methods=['GET', 'POST'])
@permission_required('导入EXCEL')
def import_contracts():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('请选择文件', 'warning')
            return redirect(url_for('import_contracts'))

        file = request.files['excel_file']
        if file.filename == '':
            flash('请选择文件', 'warning')
            return redirect(url_for('import_contracts'))

        # 新增：获取当前用户的租户ID
        customer_id = get_current_customer_id()

        file_bytes = file.read()

        try:
            df = pd.read_excel(io.BytesIO(file_bytes))

            required_cols = ['客户名称', '项目名称', '合同总价']
            missing_cols = [c for c in required_cols if c not in df.columns]
            if missing_cols:
                flash(f'Excel 格式不正确，缺少必填列：{", ".join(missing_cols)}。'
                      f'请检查列名后重新上传。当前识别到的列：{", ".join(df.columns.tolist())}', 'warning')
                return redirect(url_for('import_contracts'))

            # 合同级字段向下填充（处理Excel合并单元格导致的NaN）
            fill_cols = ['合同编号', '客户名称', '项目名称', '合同总价', '签订日期',
                         '合同类型', '业务类型', '项目负责人', '销售人员', '状态', '发票税率',
                         '已收付款', '未收付款', '已开票', '未开票']
            for col in fill_cols:
                if col in df.columns:
                    df[col] = df[col].ffill()

            # 过滤全空行
            df = df[~(df['客户名称'].isna() & df['项目名称'].isna())]

            # ── 按合同编号（若有）或(客户名称+项目名称)分组，支持多产品同一合同 ──
            from collections import OrderedDict
            contract_groups = OrderedDict()
            for idx, row in df.iterrows():
                cname = str(row.get('客户名称', '')) if pd.notna(row.get('客户名称')) else ''
                pname = str(row.get('项目名称', '')) if pd.notna(row.get('项目名称')) else ''
                cnum = str(row.get('合同编号', '')) if pd.notna(row.get('合同编号', None)) else ''
                key = cnum if cnum else (cname, pname)
                if key not in contract_groups:
                    contract_groups[key] = []
                contract_groups[key].append((idx, row))

            # 重复导入检测
            duplicates = []
            for key, rows in contract_groups.items():
                first_idx, first_row = rows[0]
                cname = str(first_row.get('客户名称', '')) if pd.notna(first_row.get('客户名称')) else ''
                pname = str(first_row.get('项目名称', '')) if pd.notna(first_row.get('项目名称')) else ''
                cnum = str(first_row.get('合同编号', '')) if pd.notna(first_row.get('合同编号', None)) else ''
                total_val = first_row.get('合同总价', 0)
                total_val = float(total_val) if pd.notna(total_val) else 0

                if cnum:
                    q = Contract.query.filter_by(contract_number=cnum)
                else:
                    q = Contract.query.filter_by(customer_name=cname, project_name=pname, total_price=total_val)
                if customer_id is not None:
                    q = q.filter_by(customer_id=customer_id)
                exists = q.first()

                if exists:
                    label = cnum if cnum else f"{cname} / {pname}"
                    duplicates.append(f"{label} / ¥{total_val}（共{len(rows)}个产品）")

            if duplicates:
                return render_template('import.html', duplicates=duplicates)

            count = 0
            errors = []
            for key, rows in contract_groups.items():
                try:
                    # 取第一行作为合同级别数据
                    first_idx, first_row = rows[0]
                    cname = str(first_row.get('客户名称', '')) if pd.notna(first_row.get('客户名称')) else ''
                    pname = str(first_row.get('项目名称', '')) if pd.notna(first_row.get('项目名称')) else ''
                    cnum = str(first_row.get('合同编号', '')) if pd.notna(first_row.get('合同编号', None)) else ''
                    total_val = first_row.get('合同总价', 0)
                    if pd.isna(total_val):
                        total_val = 0
                    staff_val = first_row.get('项目负责人') if '项目负责人' in df.columns else first_row.get('项目人员')
                    business_type_val = str(first_row.get('业务类型', '销售')) if pd.notna(first_row.get('业务类型', None)) else '销售'
                    signing_date_val = None
                    raw_date = first_row.get('签订日期', None)
                    if raw_date is not None and pd.notna(raw_date):
                        try:
                            signing_date_val = pd.to_datetime(raw_date).date()
                        except Exception:
                            signing_date_val = None

                    # 创建合同主记录
                    contract = Contract(
                        contract_number=cnum or None,
                        customer_name=cname or '未知客户',
                        project_name=pname or '未知项目',
                        total_price=float(total_val),
                        project_staff=str(staff_val) if staff_val is not None and pd.notna(staff_val) else None,
                        sales_staff=str(first_row.get('销售人员', '')) if pd.notna(first_row.get('销售人员', None)) else None,
                        business_type=business_type_val,
                        status=str(first_row.get('状态', '进行中')) if pd.notna(first_row.get('状态', None)) else '进行中',
                        signing_date=signing_date_val,
                        customer_id=customer_id  # 关联租户
                    )
                    db.session.add(contract)
                    db.session.flush()  # 获取 contract.id

                    # 自动同步客户信息
                    if cname and not Customer.query.filter_by(name=cname, customer_id=customer_id).first():
                        db.session.add(Customer(name=cname, customer_id=customer_id))

                    # ── 每一行对应一个产品 ──
                    for row_idx, row in rows:
                        product_name = str(row.get('产品名称', '')) if pd.notna(row.get('产品名称', None)) else None
                        if product_name:
                            cp = ContractProduct(
                                contract_id=contract.id,
                                product_name=product_name,
                                contract_type=str(row.get('合同类型', '')) if pd.notna(row.get('合同类型', None)) else None,
                                model=str(row.get('型号', '')) if pd.notna(row.get('型号', None)) else None,
                                unit=str(row.get('单位', '')) if pd.notna(row.get('单位', None)) else None,
                                quantity=float(row.get('数量', 0)) if pd.notna(row.get('数量', None)) else None,
                                unit_price=float(row.get('单价', 0)) if pd.notna(row.get('单价', None)) else None,
                                tax_rate=float(row.get('发票税率', 0)) if pd.notna(row.get('发票税率', None)) else None
                            )
                            # 计算小计
                            if cp.quantity and cp.unit_price:
                                cp.subtotal = cp.quantity * cp.unit_price
                            db.session.add(cp)

                            # 新增：同步产品到产品管理表
                            if not Product.query.filter_by(name=product_name, customer_id=customer_id).first():
                                model_val = str(row.get('型号', '')) if pd.notna(row.get('型号', None)) else None
                                unit_val = str(row.get('单位', '')) if pd.notna(row.get('单位', None)) else None
                                tax_val = float(row.get('发票税率', 0)) if pd.notna(row.get('发票税率', None)) else None
                                db.session.add(Product(
                                    name=product_name,
                                    category='其他',
                                    model=model_val or None,
                                    unit=unit_val or None,
                                    tax_rate=tax_val,
                                    customer_id=customer_id
                                ))

                    # 新增：处理已收付款和未收付款
                    paid_amount = first_row.get('已收付款', 0)
                    if pd.notna(paid_amount) and float(paid_amount) > 0:
                        payment = Payment(
                            contract_id=contract.id,
                            amount=float(paid_amount),
                            payment_date=signing_date_val or datetime.now().date(),
                            payment_type='导入',
                            note='Excel导入时的已收付款'
                        )
                        db.session.add(payment)

                    # 新增：处理已开票和未开票
                    invoiced_amount = first_row.get('已开票', 0)
                    if pd.notna(invoiced_amount) and float(invoiced_amount) > 0:
                        invoice = Invoice(
                            contract_id=contract.id,
                            amount=float(invoiced_amount),
                            received_date=signing_date_val or datetime.now().date(),
                            invoice_status='已开具',
                            invoice_type='普票',
                            note='Excel导入时的已开票'
                        )
                        db.session.add(invoice)

                    count += 1
                except Exception as row_err:
                    errors.append(f"{cname}/{pname}: {str(row_err)}")

            db.session.commit()
            msg = f'成功导入 {count} 条合同记录'
            if errors:
                msg += f'，{len(errors)} 条跳过：' + '；'.join(errors[:3])
            flash(msg, 'success')
            return redirect(url_for('index'))
        except Exception as e:
            db.session.rollback()
            flash(f'导入失败: {str(e)}', 'warning')
            return redirect(url_for('import_contracts'))

    return render_template('import.html')


@app.route('/download/<filename>')
@login_required
def download_file(filename):
    user_id = session.get('user_id')

    # 检查下载权限：有任何 '下载-xxx' 权限或旧格式 '下载' 权限即可
    permissions = get_user_function_permissions(user_id)
    if permissions != 'all':
        has_download_perm = '下载' in permissions or any(p.startswith('下载-') for p in permissions)
        if not has_download_perm:
            flash('你没有此项权限，请与管理员联系', 'warning')
            return redirect(url_for('index'))

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        flash('文件不存在或已被删除', 'warning')
        return redirect(request.referrer or url_for('index'))

    # 根据文件扩展名决定是否在浏览器中显示
    import mimetypes
    mimetype, _ = mimetypes.guess_type(filename)

    # 图片和PDF在浏览器中显示，其他文件下载
    if mimetype and (mimetype.startswith('image/') or mimetype == 'application/pdf'):
        return send_file(file_path, mimetype=mimetype)
    else:
        return send_file(file_path, as_attachment=True)


# 系统配置（公司图标/名称）
@app.route('/sysconfig', methods=['GET', 'POST'])
@login_required
def sysconfig():
    if session.get('role') != '超级管理员':
        flash('权限不足', 'warning')
        return redirect(url_for('index'))

    if request.method == 'POST':
        # superadmin 保存到全局配置，租户超管保存到租户配置
        if session.get('username') == 'superadmin':
            # superadmin 保存全局配置
            for key in ['system_name', 'company_name', 'company_logo']:
                val = request.form.get(key, '').strip()
                cfg = SysConfig.query.filter_by(key=key).first()
                if cfg:
                    cfg.value = val
                else:
                    db.session.add(SysConfig(key=key, value=val))
            if 'logo_file' in request.files:
                f = request.files['logo_file']
                if f.filename:
                    ext = os.path.splitext(f.filename)[1]
                    logo_filename = f'company_logo{ext}'
                    f.save(os.path.join('static', logo_filename))
                    cfg = SysConfig.query.filter_by(key='company_logo_file').first()
                    if cfg:
                        cfg.value = logo_filename
                    else:
                        db.session.add(SysConfig(key='company_logo_file', value=logo_filename))
        else:
            # 租户超管保存到租户配置
            user = db.session.get(User, session['user_id'])
            if user and user.customer_id:
                tenant = db.session.get(TenantCustomer, user.customer_id)
                if tenant:
                    tenant.system_name = request.form.get('system_name', '').strip()
                    tenant.company_name = request.form.get('company_name', '').strip()
                    if 'logo_file' in request.files:
                        f = request.files['logo_file']
                        if f.filename:
                            ext = os.path.splitext(f.filename)[1]
                            logo_filename = f'tenant_{tenant.id}_logo{ext}'
                            f.save(os.path.join('static', logo_filename))
                            tenant.logo_file = logo_filename
        db.session.commit()
        flash('配置保存成功', 'success')
        return redirect(url_for('sysconfig'))

    # 读取配置：superadmin 读全局，租户超管读租户配置
    configs = {}
    if session.get('username') == 'superadmin':
        configs = {c.key: c.value for c in SysConfig.query.all()}
    else:
        user = db.session.get(User, session['user_id'])
        if user and user.customer_id:
            tenant = db.session.get(TenantCustomer, user.customer_id)
            if tenant:
                configs['system_name'] = tenant.system_name or ''
                configs['company_name'] = tenant.company_name or ''
                configs['company_logo_file'] = tenant.logo_file or ''
    return render_template('sysconfig.html', configs=configs)


@app.route('/preview/<filename>')
@login_required
def preview_file(filename):
    user_id = session['user_id']
    if not has_permission(user_id, '查阅'):
        flash('权限不足', 'warning')
        return redirect(url_for('index'))
    from flask import Response
    import mimetypes
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    mimetype = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
    with open(file_path, 'rb') as f:
        response = Response(f.read(), mimetype=mimetype)
        response.headers['Content-Disposition'] = 'inline'
        return response


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
